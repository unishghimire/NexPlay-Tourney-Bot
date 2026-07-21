"""
NexPlay Tournament Bot v4.0 — Multi-server Discord tournament management.
Author: Unish Ghimire / NexPlay ORG | Python 3.11+

Architecture: One bot process → N servers, isolated by guild_id in all DB records.
Security: All secrets from env. Subscription gate on write commands. is_staff() role check.
"""

import discord
from discord.ext import commands
from discord import app_commands
import aiohttp
import os
import asyncio
import io
import re
import openpyxl
import yt_dlp
import random
from collections import deque
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from dotenv import load_dotenv

# Load .env file if present (local dev). Render injects vars directly.
load_dotenv()

BOT_TOKEN   = os.environ.get("DISCORD_BOT_TOKEN", "")
HOME_GUILD  = int(os.environ.get("DISCORD_GUILD_ID", "0"))   # owner's server only
SVC_TOKEN   = os.environ.get("BASE44_SERVICE_TOKEN", "")
APP_ID      = os.environ.get("APP_ID", "6a5226b5047f5c59d961130e")

BASE44_API  = "https://" + APP_ID + ".base44.app/api/apps/" + APP_ID + "/entities"
DISCORD_API = "https://discord.com/api/v10"

# ── Role names that count as "staff" in any server ────────
STAFF_ROLE_NAMES = {
    "NexPlay Owner", "Tournament Host", "Admin", "Moderator", "NexPlay Admin",
    "Owner", "Co-Owner", "Manager", "Staff",
}


# ── Status emoji map ───────────────────────────────────────
STATUS_EMOJI = {
    "registration_open":   "🟢",
    "registration_closed": "🔒",
    "groups_generated":    "🎯",
    "scheduled":           "📅",
    "in_progress":         "🔥",
    "completed":           "✅",
    "cancelled":           "❌",
}


# ────────────────────────────────────────────────────────────
#  PLAN FEATURE GATES
#  Keys match features checked throughout the bot.
#  Plans: trial / starter / pro / elite
# ────────────────────────────────────────────────────────────
PLAN_FEATURES = {
    # (plan_name_lower) → set of unlocked feature keys
    # meme_post is available on ALL plans — every server gets trending memes
    "trial":   {"create_tournament", "register", "groups", "results", "export_csv", "meme_post"},
    "starter": {"create_tournament", "register", "groups", "results", "export_csv", "meme_post",
                "schedule_post", "edit_tournament", "roadmap_post", "standings_post"},
    "pro":     {"create_tournament", "register", "groups", "results", "export_csv", "meme_post",
                "schedule_post", "edit_tournament", "roadmap_post", "standings_post",
                "ai_gfx_poster", "ai_gfx_groups", "ai_gfx_results", "ai_gfx_schedule",
                "auto_announce", "daily_report"},
    "elite":   {"create_tournament", "register", "groups", "results", "export_csv",
                "schedule_post", "edit_tournament", "roadmap_post", "standings_post",
                "ai_gfx_poster", "ai_gfx_groups", "ai_gfx_results", "ai_gfx_schedule",
                "auto_announce", "daily_report",
                "ai_support", "host_game", "suggest_improvement", "meme_post",
                "welcome_message", "advanced_analytics"},
}

PLAN_UPGRADE_MSG = {
    "ai_gfx_poster":     "🎨 AI GFX posters require **Pro** or **Elite** plan.",
    "ai_gfx_groups":     "🎨 AI GFX group draw graphics require **Pro** or **Elite** plan.",
    "ai_gfx_results":    "🎨 AI GFX result cards require **Pro** or **Elite** plan.",
    "ai_gfx_schedule":   "🎨 AI GFX schedule cards require **Pro** or **Elite** plan.",
    "auto_announce":     "📢 Auto-announce features require **Pro** or **Elite** plan.",
    "daily_report":      "📊 Daily Excel reports require **Pro** or **Elite** plan.",
    "ai_support":        "🤖 AI Support agent requires **Elite** plan.",
    "host_game":         "🎮 Mini-game hosting requires **Elite** plan.",
    "suggest_improvement":"💡 AI growth advisor requires **Elite** plan.",
    "meme_post":         "😂 Meme/clip sharing requires **Elite** plan.",
    "welcome_message":   "👋 Automated welcome messages require **Elite** plan.",
    "edit_tournament":   "✏️ Tournament editing requires **Starter** or higher plan.",
    "schedule_post":     "📅 Schedule posting requires **Starter** or higher plan.",
    "roadmap_post":      "🗺️ Roadmap posting requires **Starter** or higher plan.",
    "standings_post":    "🏅 Standings posting requires **Starter** or higher plan.",
}

async def check_feature(guild_id: str, feature: str, interaction: discord.Interaction = None) -> tuple[bool, str]:
    """
    Returns (allowed: bool, reason: str).
    If interaction is provided, automatically replies with an ephemeral error on deny.
    """
    rec = await get_server_record(guild_id)
    if not rec:
        msg = "❌ Server not registered. Run `/setup` first."
        if interaction:
            try:
                await interaction.followup.send(embed=err_e(msg), ephemeral=True)
            except:
                pass
        return False, msg

    plan    = (rec.get("plan_name") or rec.get("subscription_status") or "trial").lower()
    status  = rec.get("subscription_status", "trial").lower()

    # Normalize plan names
    if plan in ("free trial", "free_trial"): plan = "trial"
    if status in ("expired", "cancelled"):   plan = "trial"

    allowed_features = PLAN_FEATURES.get(plan, PLAN_FEATURES["trial"])
    if feature in allowed_features:
        return True, ""

    upgrade_msg = PLAN_UPGRADE_MSG.get(feature, f"❌ Feature `{feature}` not available on **{plan.title()}** plan.")
    full_msg = (
        upgrade_msg + "\n\n"
        "**Upgrade at:** https://nexplay-server-portal.vercel.app/subscription\n"
        f"Current plan: **{plan.title()}**"
    )
    if interaction:
        try:
            await interaction.followup.send(embed=err_e(full_msg), ephemeral=True)
        except:
            pass
    return False, full_msg

async def _staff_feature_gate(interaction: discord.Interaction, feature: str) -> bool:
    """Combined staff check + defer + feature gate. Returns True if allowed."""
    if not is_staff(interaction.user):
        await interaction.response.send_message(embed=err_e("❌ Staff only!"), ephemeral=True)
        return False
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), feature)
    if not ok:
        await interaction.followup.send(embed=err_e(msg), ephemeral=True)
        return False
    return True

# Shared HTTP timeout
HTTP_TIMEOUT = aiohttp.ClientTimeout(total=10)

def log(msg: str):
    """Simple print-based logger with flush for Railway."""
    print(msg, flush=True)


class NexPlayBot(commands.Bot):
    def __init__(self):
        intents = discord.Intents.default()
        intents.message_content = True
        intents.members = True
        intents.presences = True
        intents.guilds = True
        super().__init__(command_prefix="!", intents=intents)
        self.http_session: aiohttp.ClientSession | None = None

    async def setup_hook(self):
        self.http_session = aiohttp.ClientSession(timeout=HTTP_TIMEOUT)
        await self.tree.sync()
        print("[NexPlay] Global slash commands synced.")
        asyncio.create_task(self._daily_log_scheduler())
        asyncio.create_task(auto_meme_loop())
        asyncio.create_task(_247_health_check())

    async def _daily_log_scheduler(self):
        """Send daily Excel log at midnight (00:00 NPT = 18:15 UTC prev day).
           We target 00:00 NPT = UTC 18:15 = offset -5h45m from NPT.
           We compute seconds until next 18:15 UTC and sleep until then.
        """
        print("[NexPlay] Daily log scheduler started.", flush=True)
        while True:
            try:
                now_utc = datetime.now(timezone.utc)
                # Target: 18:15 UTC daily (= midnight NPT)
                target = now_utc.replace(hour=18, minute=15, second=0, microsecond=0)
                if now_utc >= target:
                    target = target.replace(day=target.day + 1)
                wait_secs = (target - now_utc).total_seconds()
                print(f"[NexPlay] Daily log fires in {wait_secs/3600:.1f}h", flush=True)
                await asyncio.sleep(wait_secs)
                await send_daily_logs(self)
            except Exception as e:
                print(f"[NexPlay] Daily log scheduler error: {e}", flush=True)
                await asyncio.sleep(3600)  # Retry in 1h on error

    async def close(self):
        if self.http_session:
            await self.http_session.close()
        await super().close()


bot  = NexPlayBot()
tree = bot.tree


MEME_SUBREDDITS = [
    "dankmemes", "gaming", "freefire", "PUBGMobile", "FreeFireBattlegrounds",
    "memes", "Minecraft", "IndianGaming", "gamingmemes", "ProgrammerHumor",
    "okbuddyretard", "wholesomememes", "PewdiepieSubmissions", "aww",
]

# Per-guild meme state tracking
_last_meme_url: dict[int, str] = {}     # guild_id → last posted meme URL
_meme_channel_cache: dict[int, int] = {} # guild_id → channel_id (cached)
MEME_INTERVAL   = 15 * 60   # post every 15 minutes

async def fetch_reddit_meme(subreddit: str = "dankmemes", exclude_url: str = "") -> dict | None:
    """Fetch a trending meme via meme-api.com.
    Tries /gimme/<subreddit> (returns top/hot post), falls back across subreddits.
    Skips NSFW, spoilers, and the last posted URL for this guild."""

    sub_list = MEME_SUBREDDITS if subreddit == "dankmemes" else [subreddit]
    # Try up to 3 different subreddits
    for attempt_sub in random.sample(sub_list, min(3, len(sub_list))):
        try:
            url = f"https://meme-api.com/gimme/{attempt_sub}"
            async with bot.http_session.get(
                url,
                headers={"User-Agent": "NexPlayBot/4.0"},
                timeout=aiohttp.ClientTimeout(total=10)
            ) as r:
                if r.status != 200:
                    continue
                data = await r.json(content_type=None)
                if not isinstance(data, dict) or "url" not in data:
                    continue
                img_url = data.get("url", "")
                if not img_url or img_url == exclude_url:
                    continue
                # Skip NSFW and spoilers
                if data.get("nsfw", False) or data.get("spoiler", False):
                    continue
                return {
                    "title":     data.get("title", "")[:256],
                    "url":       img_url,
                    "permalink": data.get("postLink", ""),
                    "score":     data.get("ups", 0),
                    "subreddit": data.get("subreddit", attempt_sub),
                }
        except Exception as e:
            print(f"[NexPlay] meme fetch error r/{attempt_sub}: {e}", flush=True)
            continue

    # Fallback: fetch a random meme (no subreddit filter)
    try:
        async with bot.http_session.get(
            "https://meme-api.com/gimme",
            headers={"User-Agent": "NexPlayBot/4.0"},
            timeout=aiohttp.ClientTimeout(total=10)
        ) as r:
            if r.status == 200:
                data = await r.json(content_type=None)
                img_url = data.get("url", "")
                if img_url and img_url != exclude_url and not data.get("nsfw", False):
                    return {
                        "title":     data.get("title", "")[:256],
                        "url":       img_url,
                        "permalink": data.get("postLink", ""),
                        "score":     data.get("ups", 0),
                        "subreddit": data.get("subreddit", "random"),
                    }
    except Exception as e:
        print(f"[NexPlay] meme fetch fallback error: {e}", flush=True)

    return None

async def auto_meme_loop():
    """Post trending memes to ALL servers every 30 minutes.
    Elite servers: every 30 min. Non-Elite: every 2 hours (4 cycles)."""
    await bot.wait_until_ready()
    print(f"[NexPlay] auto_meme_loop started — interval={MEME_INTERVAL}s (15 min, ALL servers)", flush=True)
    while not bot.is_closed():
        try:
            for guild in bot.guilds:
                try:
                    # ALL servers get memes every 15 minutes
                    ok, _ = await check_feature(str(guild.id), "meme_post")
                    if not ok:
                        continue  # Skip servers without meme_post feature

                    # Find meme channel (cached or by name hint)
                    target = guild.get_channel(_meme_channel_cache.get(guild.id, 0))
                    if target and not target.permissions_for(guild.me).send_messages: target = None
                    if not target:
                        for hint in ("meme-server", "memes", "meme", "funny", "media", "general", "chat"):
                            target = discord.utils.find(
                                lambda c, h=hint: isinstance(c, discord.TextChannel) and h in c.name.lower()
                                and c.permissions_for(guild.me).send_messages, guild.text_channels)
                            if target:
                                _meme_channel_cache[guild.id] = target.id
                                break
                    if not target: continue

                    # Pick subreddit — rotate per guild per cycle for variety
                    sub_idx = (guild.id // 1000 + int(asyncio.get_event_loop().time()) // 900) % len(MEME_SUBREDDITS)
                    sub = MEME_SUBREDDITS[sub_idx]

                    # Fetch trending meme (fetch_reddit_meme already tries 3 subreddits)
                    last_url = _last_meme_url.get(guild.id, "")
                    meme = await fetch_reddit_meme(sub, exclude_url=last_url)
                    if not meme:
                        continue

                    # Save URL to prevent repeats
                    _last_meme_url[guild.id] = meme["url"]

                    # Build trending embed
                    score = meme.get("score", 0)
                    is_hot = score > 1000
                    embed = discord.Embed(
                        title=f"🔥 {meme['title']}",
                        url=meme["permalink"],
                        color=0xFF6B35 if is_hot else 0x5865F2
                    )
                    embed.set_image(url=meme["url"])
                    footer_text = f"r/{sub} • {'🔥 HOT' if is_hot else 'Trending'} • 👍 {score:,} | NexPlay Memes"
                    embed.set_footer(text=footer_text)

                    await target.send(embed=embed)
                    print(f"[NexPlay] Meme posted → #{target.name} in {guild.name} (r/{sub}, score={score})", flush=True)

                except Exception as guild_err:
                    print(f"[NexPlay] meme loop error ({guild.name}): {guild_err}", flush=True)

        except Exception as e:
            print(f"[NexPlay] auto_meme_loop outer error: {e}", flush=True)

        await asyncio.sleep(MEME_INTERVAL)


@bot.event
async def on_member_join(member: discord.Member):
    """Welcome new members (Elite) + update server member count."""
    guild = member.guild
    # Update member count for ALL servers
    await _update_member_count(guild)
    # Welcome message (Elite only)
    ok, _ = await check_feature(str(guild.id), "welcome_message")
    if not ok:
        return
    # Find welcome channel
    welcome_ch = discord.utils.find(
        lambda c: isinstance(c, discord.TextChannel) and "welcome" in c.name.lower(),
        guild.text_channels
    )
    if not welcome_ch:
        return
    e = discord.Embed(title=f"👋 Welcome to {guild.name}, {member.display_name}!",
        description=(f"Hey {member.mention}, welcome! 🎮\n\n"
            f"**{guild.name}** — competitive esports & Free Fire tournaments.\n\n"
            f"📋 Check our channels\n🏆 Use `/register` to join a tournament\n"
            f"❓ Need help? Use `/help`\n\nGood luck and have fun! 🔥"),
        color=0x5865F2, timestamp=datetime.now(timezone.utc))
    e.set_thumbnail(url=member.display_avatar.url)
    e.set_footer(text=f"Member #{guild.member_count} • NexPlay")
    await welcome_ch.send(embed=e)


@tree.command(name="announce", description="[Host] Post a tournament announcement embed")
@app_commands.describe(
    title="Announcement title",
    message="Announcement message body",
    channel="Channel to post in (optional, defaults to current)"
)
async def cmd_announce(
    interaction: discord.Interaction,
    title: str,
    message: str,
    channel: discord.TextChannel = None
):
    if not is_staff(interaction):
        return await interaction.response.send_message(embed=err_e("Staff only."), ephemeral=True)
    await interaction.response.defer(thinking=True, ephemeral=True)
    target = channel or interaction.channel
    embed = discord.Embed(
        title=f"📢 {title}",
        description=message,
        color=0xE74C3C,
        timestamp=datetime.now(timezone.utc)
    )
    embed.set_footer(text=f"NexPlay Announcement • {interaction.guild.name}")
    await target.send(embed=embed)
    await interaction.followup.send(embed=ok_e("Announced!", f"Posted to {target.mention}"), ephemeral=True)

# ────────────────────────────────────────────────────────────
#  MULTI-STEP TOURNAMENT CREATION MODALS
# ────────────────────────────────────────────────────────────

#  TOURNAMENT CREATION — MODAL → BUTTON → MODAL FLOW
#  Discord does NOT allow opening a modal from inside another
#  modal's on_submit. Pattern: Modal → ephemeral View with
#  Next button → button opens next Modal → repeat.
#  Step data stored in _tourney_sessions[user_id].

_tourney_sessions: dict[int, dict] = {}

FMT_LABELS = {"single_elim":"Single Elimination","double_elim":"Double Elimination","round_robin":"Round Robin","battle_royale":"Battle Royale"}
STAGE_NAMES = ["Group Stage","Quarter Final","Semi Final","Grand Final","Championship"]

async def _post_tournament_info(ch_info, ch_ann, ch_reg, ch_road, ch_logo, name, game,
                                fmt_label, prize, date, time_str, max_p, t_size, g_size,
                                rounds, reg_end, nations, rules, stream, desc, poster, roadmap):
    if ch_info:
        d = (f"**🎮 Game:** {game}\n**🏆 Format:** {fmt_label}\n**💰 Prize:** {prize}\n"
             f"**📅 Date:** {date} | **⏰ Time:** {time_str}\n"
             f"**👥 Max Teams:** {max_p} | **Size:** {t_size}v{t_size}\n"
             f"**🎯 Groups:** {g_size}/group | **Rounds:** {rounds}\n"
             f"**🗓️ Reg Deadline:** {reg_end} | **🌏 Nations:** {nations}")
        if rules: d += f"\n**📜 Rules:**\n{rules}"
        if stream: d += f"\n**📺 Stream:** {stream}"
        if desc: d += f"\n{desc}"
        e = discord.Embed(title=f"📋 {name} — Tournament Info", description=d, color=0x5865F2, timestamp=datetime.now(timezone.utc))
        e.set_footer(text="NexPlay Tournament System")
        if poster: e.set_image(url=poster)
        await dpost(ch_info, e)

    if ch_ann:
        d = (f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n🎮 **Game:** {game}\n💰 **Prize:** {prize}\n"
             f"📅 **Date:** {date} | ⏰ {time_str}\n👥 **Slots:** {max_p} teams ({t_size}v{t_size})\n"
             f"🌏 **Nations:** {nations}\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        if ch_reg: d += f"\n✍️ Register → <#{ch_reg}>"
        if ch_info: d += f"\n📋 Info → <#{ch_info}>"
        if ch_road: d += f"\n🗺️ Roadmap → <#{ch_road}>"
        if stream: d += f"\n📺 Stream: **{stream}**"
        e = discord.Embed(title=f"🚨 {name.upper()} — REGISTRATION OPEN!", description=d, color=0x00FF7F, timestamp=datetime.now(timezone.utc))
        e.set_footer(text="NexPlay | Registration Open")
        if poster: e.set_image(url=poster)
        await dpost(ch_ann, e)

    if ch_reg:
        lines = "\n".join([f"Player {i+1}: @mention" for i in range(t_size)])
        d = (f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
             f"🎮 **Game:** {game}\n💰 **Prize:** {prize}\n"
             f"📅 **Date:** {date} | ⏰ {time_str}\n"
             f"👥 **Slots:** {max_p} teams ({t_size}v{t_size})\n"
             f"🗓️ **Deadline:** {reg_end}\n"
             f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
             f"@everyone **Registration OPEN!**\n"
             f"```\nTeam Name: <name>\n{lines}\n```\n"
             f"⚠️ All {t_size} players must be @mentioned.")
        if ch_logo: d += f"\n🎨 Submit logo in <#{ch_logo}> — `Team Name: <name>` + image"
        e = discord.Embed(title=f"✍️ Register for {name}", description=d, color=0x00FF7F, timestamp=datetime.now(timezone.utc))
        e.set_footer(text="NexPlay Registration")
        await dpost(ch_reg, e)

    if ch_road:
        lines = [f"**Stage {i+1}:** {STAGE_NAMES[i] if i < 5 else f'Round {i+1}'}" for i in range(min(rounds, 5))]
        e = discord.Embed(title=f"🗺️ {name} — Roadmap", description="\n".join(lines), color=0xFF6B35)
        e.set_footer(text="NexPlay Roadmap")
        if roadmap: e.set_image(url=roadmap)
        await dpost(ch_road, e)

# ── Step 1 Modal ──────────────────────────────────────────
class TournamentStep1Modal(discord.ui.Modal, title="🏆 Create Tournament (1/3) — Basics"):
    t_name  = discord.ui.TextInput(label="Tournament Name",       placeholder="NexPlay Championship S1",                 max_length=80)
    t_game  = discord.ui.TextInput(label="Game",                  placeholder="Free Fire / PUBG / Minecraft / Valorant", max_length=40)
    t_prize = discord.ui.TextInput(label="Prize Pool",            placeholder="NPR 10,000",                              max_length=60)
    t_date  = discord.ui.TextInput(label="Tournament Date",       placeholder="2026-08-01",                              max_length=30)
    t_desc  = discord.ui.TextInput(label="Description (optional)",placeholder="Open for all — register now!",            required=False, max_length=200, style=discord.TextStyle.paragraph)

    async def on_submit(self, interaction: discord.Interaction):
        _tourney_sessions[interaction.user.id] = {
            "name":        self.t_name.value.strip(),
            "game":        self.t_game.value.strip(),
            "prize_pool":  self.t_prize.value.strip(),
            "date":        self.t_date.value.strip(),
            "description": self.t_desc.value.strip(),
        }
        embed = discord.Embed(
            title="✅ Step 1 of 3 saved!",
            description=(
                f"**Name:** {self.t_name.value.strip()}\n"
                f"**Game:** {self.t_game.value.strip()}\n"
                f"**Prize:** {self.t_prize.value.strip()}\n"
                f"**Date:** {self.t_date.value.strip()}\n\n"
                "Click **Next →** to fill Format & Slots."
            ),
            color=0x5865F2
        )
        await interaction.response.send_message(embed=embed, view=Step2ButtonView(), ephemeral=True)


class Step2ButtonView(discord.ui.View):
    def __init__(self): super().__init__(timeout=300)

    @discord.ui.button(label="Next → Format & Slots", style=discord.ButtonStyle.primary, emoji="➡️")
    async def go_step2(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id not in _tourney_sessions:
            return await interaction.response.send_message("❌ Session expired. Run /create_tournament again.", ephemeral=True)
        await interaction.response.send_modal(TournamentStep2Modal())

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.danger, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        _tourney_sessions.pop(interaction.user.id, None)
        await interaction.response.edit_message(content="❌ Tournament creation cancelled.", embed=None, view=None)


# ── Step 2 Modal ──────────────────────────────────────────
class TournamentStep2Modal(discord.ui.Modal, title="🏆 Create Tournament (2/3) — Format & Slots"):
    t_max    = discord.ui.TextInput(label="Total Teams / Max Slots", placeholder="16",                                    max_length=4)
    t_tsize  = discord.ui.TextInput(label="Players Per Team",        placeholder="4  (squads) or 1 (solo)",               max_length=3)
    t_gsize  = discord.ui.TextInput(label="Teams Per Group",         placeholder="4",                                     max_length=3)
    t_rounds = discord.ui.TextInput(label="Number of Rounds",        placeholder="3  (Group → Semi → Final)",             max_length=3)
    t_format = discord.ui.TextInput(label="Match Format",            placeholder="Battle Royale / Single Elim / Round Robin", max_length=40)

    async def on_submit(self, interaction: discord.Interaction):
        uid = interaction.user.id
        if uid not in _tourney_sessions:
            return await interaction.response.send_message("❌ Session expired. Run /create_tournament again.", ephemeral=True)
        _tourney_sessions[uid].update({
            "max_players": self.t_max.value.strip(),
            "team_size":   self.t_tsize.value.strip(),
            "group_size":  self.t_gsize.value.strip(),
            "rounds":      self.t_rounds.value.strip(),
            "format":      self.t_format.value.strip(),
        })
        embed = discord.Embed(
            title="✅ Step 2 of 3 saved!",
            description=(
                f"**Teams:** {self.t_max.value.strip()}  |  **Size:** {self.t_tsize.value.strip()}v{self.t_tsize.value.strip()}\n"
                f"**Groups:** {self.t_gsize.value.strip()} teams/group  |  **Rounds:** {self.t_rounds.value.strip()}\n"
                f"**Format:** {self.t_format.value.strip()}\n\n"
                "Click **Next →** to fill Schedule & Rules (last step!)"
            ),
            color=0x5865F2
        )
        await interaction.response.send_message(embed=embed, view=Step3ButtonView(), ephemeral=True)


class Step3ButtonView(discord.ui.View):
    def __init__(self): super().__init__(timeout=300)

    @discord.ui.button(label="Next → Schedule & Rules", style=discord.ButtonStyle.primary, emoji="➡️")
    async def go_step3(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id not in _tourney_sessions:
            return await interaction.response.send_message("❌ Session expired. Run /create_tournament again.", ephemeral=True)
        await interaction.response.send_modal(TournamentStep3Modal())

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.danger, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        _tourney_sessions.pop(interaction.user.id, None)
        await interaction.response.edit_message(content="❌ Tournament creation cancelled.", embed=None, view=None)


# ── Step 3 Modal — saves to DB ────────────────────────────
class TournamentStep3Modal(discord.ui.Modal, title="🏆 Create Tournament (3/3) — Schedule & Rules"):
    t_time    = discord.ui.TextInput(label="Match Time",            placeholder="5:00 PM NPT",                           max_length=30)
    t_regend  = discord.ui.TextInput(label="Registration Deadline", placeholder="2026-07-30 or TBD", required=False,      max_length=30)
    t_rules   = discord.ui.TextInput(label="Rules / Special Notes", placeholder="No hacks. Top 2 per group advance.",    required=False, max_length=400, style=discord.TextStyle.paragraph)
    t_stream  = discord.ui.TextInput(label="Stream / YouTube",      placeholder="@nexplaygg or https://youtube.com/...", required=False, max_length=100)
    t_nations = discord.ui.TextInput(label="Eligible Nations",      placeholder="🇳🇵 🇮🇳 🇧🇩 🇵🇰  (blank = open)",          required=False, max_length=60)

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True, ephemeral=True)
        uid = interaction.user.id
        if uid not in _tourney_sessions:
            return await interaction.followup.send("❌ Session expired. Run /create_tournament again.", ephemeral=True)

        s   = _tourney_sessions.pop(uid)
        gid = str(interaction.guild.id)


        name        = s["name"]
        game        = s["game"]
        prize_pool  = s["prize_pool"]
        date        = s["date"]
        description = s.get("description", "")
        max_players = safe_int(s.get("max_players", "16"), 16)
        team_size   = safe_int(s.get("team_size", "4"),     4)
        group_size  = safe_int(s.get("group_size", "4"),    4)
        rounds      = safe_int(s.get("rounds", "3"),        3)
        fmt         = s.get("format", "Battle Royale")
        time_str    = self.t_time.value.strip()    or "TBD"
        reg_end     = self.t_regend.value.strip()  or "TBD"
        rules       = self.t_rules.value.strip()
        stream      = self.t_stream.value.strip()
        nations     = self.t_nations.value.strip() or "🇳🇵"

        # Create channels
        t_channels = await make_tournament_channels(interaction.guild, name)
        short      = t_channels.get("short_name", "tourney")
        cat_name   = t_channels.get("category_name", "🏆 Tournament")

        def ch_id(key):
            v = t_channels.get(key)
            return int(v) if v else None

        ch_ann_id  = ch_id("announcements")
        ch_reg_id  = ch_id("register")
        ch_info_id = ch_id("info")
        ch_road_id = ch_id("roadmap")
        ch_logo_id = ch_id("team-logo")

        # AI GFX (Pro+)
        can_gfx, _ = await check_feature(gid, "ai_gfx_poster")
        poster_url  = img_url(name, game, "poster", f"prize {prize_pool} date {date}") if can_gfx else ""
        roadmap_url = img_url(name, game, "roadmap") if can_gfx else ""

        # Save to DB
        rec = await b44_create("Tournament", {
            "guild_id": gid, "name": name, "game": game, "format": fmt,
            "prize_pool": prize_pool, "description": description,
            "status": "registration_open", "max_players": max_players,
            "team_size": team_size, "group_size": group_size, "rounds": rounds,
            "rules": rules, "stream_channel": stream, "eligible_nations": nations,
            "reg_deadline": reg_end, "registered_count": 0,
            "tournament_date": date, "tournament_time": time_str,
            "poster_image_url": poster_url, "roadmap_image_url": roadmap_url,
            "announcement_channel_id": str(ch_ann_id) if ch_ann_id else "",
            "registration_channel_id": str(ch_reg_id) if ch_reg_id else "",
            "short_name": short, "category_name": cat_name,
            "created_by_discord_id": str(interaction.user.id),
            "started_at": now_iso(),
        })
        tid = rec.get("id", "")

        fmt_label = FMT_LABELS.get(fmt.lower().replace(" ", "_"), fmt)

        await _post_tournament_info(ch_info_id, ch_ann_id, ch_reg_id, ch_road_id, ch_logo_id,
            name, game, fmt_label, prize_pool, date, time_str, max_players,
            team_size, group_size, rounds, reg_end, nations, rules, stream,
            description, poster_url, roadmap_url)

        # Update server tournament count
        srv = await get_server_record(gid)
        if srv:
            await b44_update("Server", srv["id"], {
                "tournaments_used": (srv.get("tournaments_used") or 0) + 1,
                "last_active": now_iso(),
            })

        # Confirm to host
        confirm = discord.Embed(
            title="🎉 Tournament Created!",
            description=(
                f"**{name}** is live!\n\n"
                f"📁 Category: **{cat_name}**\n"
                + (f"📢 Announcements: <#{ch_ann_id}>\n" if ch_ann_id else "")
                + (f"✍️ Registration: <#{ch_reg_id}>\n" if ch_reg_id else "")
                + (f"🖼️ Logo channel: <#{ch_logo_id}>\n" if ch_logo_id else "")
                + f"\n`ID: {tid}`\nNext: `/generate_groups` after reg closes."
            ),
            color=0x00FF7F
        )
        confirm.set_footer(text="NexPlay Tournament System")
        await interaction.followup.send(embed=confirm, ephemeral=True)


class TournamentEditModal(discord.ui.Modal, title="✏️ Edit Tournament"):
    t_prize     = discord.ui.TextInput(label="Prize Pool",         max_length=60)
    t_date      = discord.ui.TextInput(label="Tournament Date",    max_length=30)
    t_time      = discord.ui.TextInput(label="Match Time",         max_length=30)
    t_rules     = discord.ui.TextInput(label="Rules / Notes",      max_length=400, required=False, style=discord.TextStyle.paragraph)
    t_stream    = discord.ui.TextInput(label="Stream Channel",     max_length=100, required=False)

    def __init__(self, tournament: dict):
        super().__init__()
        self.tournament = tournament
        # Pre-fill with existing values
        self.t_prize.default  = tournament.get("prize_pool", "")
        self.t_date.default   = tournament.get("tournament_date", "")
        self.t_time.default   = tournament.get("tournament_time", "TBD")
        self.t_rules.default  = tournament.get("rules", "")
        self.t_stream.default = tournament.get("stream_channel", "")

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True, ephemeral=True)
        t = self.tournament

        updates = {
            "prize_pool":        self.t_prize.value.strip(),
            "tournament_date":   self.t_date.value.strip(),
            "tournament_time":   self.t_time.value.strip(),
            "rules":             self.t_rules.value.strip(),
            "stream_channel":    self.t_stream.value.strip(),
        }
        await b44_update("Tournament", t["id"], updates)

        # Refresh #info channel embed
        gid   = str(interaction.guild.id)
        t.update(updates)
        ch_info_id = None
        short = t.get("short_name", "")
        if short:
            info_ch = discord.utils.get(interaction.guild.text_channels, name=short + "-info")
            if info_ch: ch_info_id = info_ch.id

        if ch_info_id:
            t.update(updates)
            d = (f"**🎮 Game:** {t.get('game','')}\n**💰 Prize:** {updates['prize_pool']}\n"
                 f"**📅 Date:** {updates['tournament_date']} | ⏰ {updates['tournament_time']}\n"
                 f"**👥 Slots:** {t.get('max_players','')} | Teams: {t.get('team_size','')}v{t.get('team_size','')}\n"
                 f"**🎯 Group:** {t.get('group_size','')}/group | Rounds: {t.get('rounds','')}\n"
                 f"**🌏 Nations:** {t.get('eligible_nations','🇳🇵')}")
            if updates['rules']: d += f"\n**📜 Rules:**\n{updates['rules']}"
            if updates['stream_channel']: d += f"\n**📺 Stream:** {updates['stream_channel']}"
            await dpost(ch_info_id, discord.Embed(
                title=f"📋 {t['name']} — Updated", description=d, color=0xFFA500,
                timestamp=datetime.now(timezone.utc)).set_footer(text="NexPlay | Updated"))

        await interaction.followup.send(embed=ok_e("Tournament Updated!", f"**{t['name']}** details have been updated and #info channel refreshed."), ephemeral=True)


class TournamentEditSlotView(discord.ui.View):
    """Second edit modal covering slots/format/nations."""
    def __init__(self, tournament: dict):
        super().__init__(timeout=120)
        self.tournament = tournament

    @discord.ui.button(label="Edit Basic Info", style=discord.ButtonStyle.primary, emoji="✏️")
    async def edit_basic(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(TournamentEditModal(self.tournament))

    @discord.ui.button(label="Edit Slots & Nations", style=discord.ButtonStyle.secondary, emoji="🎯")
    async def edit_slots(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(TournamentEditSlotsModal(self.tournament))

    @discord.ui.button(label="Change Status", style=discord.ButtonStyle.danger, emoji="🔄")
    async def change_status(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(TournamentStatusModal(self.tournament))


class TournamentEditSlotsModal(discord.ui.Modal, title="✏️ Edit Tournament — Slots & Nations"):
    t_max     = discord.ui.TextInput(label="Max Teams",          max_length=4)
    t_tsize   = discord.ui.TextInput(label="Team Size",          max_length=3)
    t_gsize   = discord.ui.TextInput(label="Teams Per Group",    max_length=3)
    t_rounds  = discord.ui.TextInput(label="Number of Rounds",   max_length=3)
    t_nations = discord.ui.TextInput(label="Eligible Nations",   max_length=60, required=False)

    def __init__(self, tournament: dict):
        super().__init__()
        self.tournament = tournament
        self.t_max.default     = str(tournament.get("max_players", ""))
        self.t_tsize.default   = str(tournament.get("team_size", ""))
        self.t_gsize.default   = str(tournament.get("group_size", ""))
        self.t_rounds.default  = str(tournament.get("rounds", ""))
        self.t_nations.default = tournament.get("eligible_nations", "🇳🇵")

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True, ephemeral=True)
        t = self.tournament
        updates = {
            "max_players":      safe_int(self.t_max.value, t.get("max_players", 16)),
            "team_size":        safe_int(self.t_tsize.value, t.get("team_size", 4)),
            "group_size":       safe_int(self.t_gsize.value, t.get("group_size", 4)),
            "rounds":           safe_int(self.t_rounds.value, t.get("rounds", 3)),
            "eligible_nations": self.t_nations.value.strip() or "🇳🇵",
        }
        await b44_update("Tournament", t["id"], updates)
        await interaction.followup.send(embed=ok_e("Slots Updated!", "Tournament structure updated successfully."), ephemeral=True)


class TournamentStatusModal(discord.ui.Modal, title="🔄 Change Tournament Status"):
    t_status = discord.ui.TextInput(
        label="New Status",
        placeholder="registration_open / registration_closed / in_progress / completed / cancelled",
        max_length=30
    )

    def __init__(self, tournament: dict):
        super().__init__()
        self.tournament = tournament
        self.t_status.default = tournament.get("status", "registration_open")

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True, ephemeral=True)
        valid = {"registration_open", "registration_closed", "in_progress", "completed", "cancelled"}
        new_status = self.t_status.value.strip().lower()
        if new_status not in valid:
            return await interaction.followup.send(embed=err_e(f"Invalid status. Choose from: {', '.join(valid)}"), ephemeral=True)
        await b44_update("Tournament", self.tournament["id"], {"status": new_status})
        await interaction.followup.send(embed=ok_e("Status Updated!", f"Tournament status → **{new_status}**"), ephemeral=True)


@bot.event
async def on_ready():
    """Fires when bot connects. Clears support locks for fresh session."""
    global _replied_users
    _replied_users = {}
    print(f"[NexPlay] ✅ {bot.user} online — {len(bot.guilds)} server(s)", flush=True)
    for g in bot.guilds:
        print(f"[NexPlay]   • {g.name} ({g.id}) — {g.member_count} members", flush=True)
    if SVC_TOKEN:
        try: print(f"[NexPlay] DB OK — {len(await b44_list('Server'))} server(s)", flush=True)
        except Exception as e: print(f"[NexPlay] ⚠️ DB: {e}", flush=True)
    else:
        print("[NexPlay] ⚠️ BASE44_SERVICE_TOKEN not set!", flush=True)
    try: print(f"[NexPlay] Synced {len(await bot.tree.sync())} commands.", flush=True)
    except Exception as e: print(f"[NexPlay] ⚠️ Sync: {e}", flush=True)


def _b44_headers() -> dict:
    return {"Authorization": "Bearer " + SVC_TOKEN, "Content-Type": "application/json"}

async def b44_list(entity: str, filters: dict | None = None) -> list:
    url = BASE44_API + "/" + entity
    try:
        if not SVC_TOKEN:
            print("[b44_list] ⚠️ SVC_TOKEN is empty! Base44 API calls will fail.", flush=True)
            return []
        async with bot.http_session.get(url, headers=_b44_headers()) as r:
            if r.status != 200:
                err_body = await r.text()
                print(f"[b44_list] ⚠️ {entity} HTTP {r.status}: {err_body[:200]}", flush=True)
                return []
            data = await r.json()
            if not isinstance(data, list):
                print(f"[b44_list] ⚠️ {entity} returned non-list: {type(data).__name__}", flush=True)
                return []
            if filters:
                for k, v in filters.items():
                    data = [x for x in data if x.get(k) == v]
            return data
    except Exception as e:
        print(f"[b44_list] ❌ {entity} error: {e}", flush=True)
        return []


async def b44_create(entity: str, payload: dict) -> dict:
    url = BASE44_API + "/" + entity
    if not SVC_TOKEN:
        print("[b44_create] ⚠️ SVC_TOKEN is empty! Base44 API calls will fail.", flush=True)
        return {}
    try:
        async with bot.http_session.post(url, json=payload, headers=_b44_headers()) as r:
            if r.status in (200, 201):
                return await r.json()
            err_body = await r.text()
            print(f"[b44_create] ⚠️ {entity} HTTP {r.status}: {err_body[:200]}", flush=True)
            return {}
    except Exception as e:
        print(f"[b44_create] ❌ {entity} error: {e}", flush=True)
        return {}


async def b44_update(entity: str, record_id: str, payload: dict) -> dict:
    url = BASE44_API + "/" + entity + "/" + record_id
    try:
        async with bot.http_session.put(url, json=payload, headers=_b44_headers()) as r:
            if r.status in (200, 201):
                return await r.json()
            err_body = await r.text()
            print(f"[b44_update] ⚠️ {entity}/{record_id} HTTP {r.status}: {err_body[:200]}", flush=True)
            return {}
    except Exception as e:
        print(f"[b44_update] ❌ {entity} error: {e}", flush=True)
        return {}


async def b44_delete(entity: str, record_id: str) -> bool:
    url = BASE44_API + "/" + entity + "/" + record_id
    try:
        async with bot.http_session.delete(url, headers=_b44_headers()) as r:
            if r.status in (200, 204):
                return True
            err_body = await r.text()
            print(f"[b44_delete] ⚠️ {entity}/{record_id} HTTP {r.status}: {err_body[:200]}", flush=True)
            return False
    except Exception as e:
        print(f"[b44_delete] ❌ {entity} error: {e}", flush=True)
        return False

async def _b44_delete_by_tournament(entity: str, tournament_id: str) -> int:
    """Delete all records of an entity matching tournament_id. Returns count."""
    count = 0
    for rec in await b44_list(entity, {"tournament_id": tournament_id}):
        if rec.get("id") and await b44_delete(entity, rec["id"]):
            count += 1
    return count


#  SUBSCRIPTION GATE
#  Every staff command calls this first. If the server isn't
#  registered or has expired, the command is blocked.
async def get_server_record(guild_id: str) -> dict | None:
    servers = await b44_list("Server", {"guild_id": guild_id})
    return servers[0] if servers else None

async def _update_member_count(guild: discord.Guild):
    """Update server member count + last_active in DB."""
    srv = await get_server_record(str(guild.id))
    if srv:
        await b44_update("Server", srv["id"], {"member_count": guild.member_count or 0, "last_active": now_iso()})

async def register_server(guild: discord.Guild) -> dict:
    """Register a new server as Free Trial. Returns the created record."""
    owner = guild.owner
    return await b44_create("Server", {
        "guild_id":            str(guild.id),
        "guild_name":          guild.name,
        "owner_id":            str(owner.id) if owner else "",
        "owner_name":          owner.display_name if owner else "Unknown",
        "plan_name":           "Free Trial",
        "subscription_status": "trial",
        "tournaments_used":    0,
        "tournament_limit":    3,
        "member_count":        guild.member_count or 0,
        "last_active":         now_iso(),
    })


async def is_allowed(guild_id: str, guild: discord.Guild = None) -> tuple[bool, str]:
    """Returns (allowed, reason). Auto-registers new servers as Free Trial."""
    rec = await get_server_record(guild_id)
    if not rec:
        if guild:
            owner = guild.owner
            rec = await b44_create("Server", {
                "guild_id":            guild_id,
                "guild_name":          guild.name,
                "owner_id":            str(owner.id) if owner else "",
                "owner_name":          owner.display_name if owner else "Unknown",
                "plan_name":           "Free Trial",
                "subscription_status": "trial",
                "tournaments_used":    0,
                "tournament_limit":    3,
                "member_count":        guild.member_count or 0,
                "last_active":         now_iso(),
            })
            print(f"[NexPlay] Auto-registered {guild.name} via is_allowed fallback", flush=True)
        else:
            return False, "This server is not registered with NexPlay. Please re-invite the bot."
    status = rec.get("subscription_status", "trial")
    if status == "banned":
        return False, "This server has been banned from NexPlay. Contact support."
    if status == "inactive":
        return False, "This server was previously removed. Run `/setup` to reactivate your account. Your previous data has been preserved."
    if status in ("active", "trial"):
        return True, ""
    return False, "This server's NexPlay subscription has expired.\nRenew at: https://nexplay-server-portal.vercel.app/subscription\nPlans from NPR 99/mo (Starter) · NPR 299/mo (Pro AI) · NPR 399/mo (Elite)"


#  DYNAMIC CHANNEL RESOLVER
#  Finds channels by name — works in ANY server.
#  Falls back to first available channel if not found.
CHANNEL_NAMES = {
    "announcements": ["tourney-announcements", "tournament-announcements", "announcements", "general"],
    "registration":  ["tourney-registration", "tournament-registration", "registration", "general"],
    "brackets":      ["brackets-results", "brackets", "results", "tournament-results", "general"],
    "champions":     ["hall-of-champions", "champions", "winners", "general"],
    "support":       ["support-ticket", "support", "help", "general"],
    "staff":         ["mod-log", "staff-log", "staff", "moderators", "admin-log"],
    "rules":         ["tourney-rules", "rules", "tournament-rules", "general"],
}


def resolve_channel(guild: discord.Guild, key: str) -> discord.TextChannel | None:
    """Find a channel by trying a list of common names. No fallback — returns None if not found."""
    for name in CHANNEL_NAMES.get(key, []):
        ch = discord.utils.get(guild.text_channels, name=name)
        if ch:
            return ch
    return None


async def get_or_create_channels(guild: discord.Guild) -> dict:
    """
    Ensure required channels exist. Returns a dict of channel IDs.
    Creates missing channels under a 'NexPlay Tournaments' category.
    Only creates a channel if it doesn't already exist — no duplicates.
    """
    required = {
        "announcements": "tourney-announcements",
        "registration":  "tourney-registration",
        "brackets":      "brackets-results",
        "champions":     "hall-of-champions",
        "support":       "support-ticket",
        "staff":         "mod-log",
    }
    result = {}
    category = discord.utils.get(guild.categories, name="NexPlay Tournaments")
    if not category:
        try:
            category = await guild.create_category("NexPlay Tournaments")
        except Exception:
            category = None

    for key, ch_name in required.items():
        ch = resolve_channel(guild, key)
        if not ch:
            # Double-check by exact name to avoid duplicates
            existing = discord.utils.get(guild.text_channels, name=ch_name)
            if existing:
                ch = existing
            else:
                try:
                    ch = await guild.create_text_channel(ch_name, category=category)
                except Exception:
                    ch = None
        result[key] = ch.id if ch else None
    return result


async def dpost(channel_id: int, embed: discord.Embed, content: str = "") -> dict:
    if not channel_id:
        return {}
    url     = DISCORD_API + "/channels/" + str(channel_id) + "/messages"
    headers = {"Authorization": "Bot " + BOT_TOKEN, "Content-Type": "application/json"}
    body    = {"embeds": [embed.to_dict()]}
    if content:
        body["content"] = content
    try:
        async with bot.http_session.post(url, json=body, headers=headers) as r:
            return await r.json()
    except Exception as e:
        print("[dpost] " + str(e))
        return {}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def now_ts() -> int:
    return int(datetime.now(timezone.utc).timestamp())

def safe_int(v, d=16):
    """Safe int conversion with default fallback."""
    try: return int(v)
    except: return d

def is_staff(member: discord.Member) -> bool:
    if member.guild.owner_id == member.id:
        return True
    for role in member.roles:
        # Strip emoji prefixes (any non-alphanumeric leading chars)
        clean = re.sub(r"^[^A-Za-z]+", "", role.name).strip()
        if clean in STAFF_ROLE_NAMES:
            return True
    return False

def err_e(msg: str) -> discord.Embed:
    e = discord.Embed(description="❌  " + msg, color=0xFF4444)
    e.set_footer(text="NexPlay")
    return e

def ok_e(title: str, desc: str, color: int = 0x00FF7F) -> discord.Embed:
    e = discord.Embed(title=title, description=desc, color=color)
    e.set_footer(text="NexPlay Tournament System")
    e.timestamp = datetime.now(timezone.utc)
    return e

async def _reject_msg(message: discord.Message, title: str, desc: str, delay: int = 15):
    """Reply with an error embed, react ❌, then delete both after delay."""
    await message.add_reaction("❌")
    err = await message.reply(
        embed=discord.Embed(title=title, description=desc, color=0xFF4444),
        mention_author=True
    )
    await asyncio.sleep(delay)
    try:
        await err.delete()
        await message.delete()
    except:
        pass

def img_url(t_name: str, game: str, kind: str, extra: str = "") -> str:
    """Generate a fixed-size tournament image URL via Pollinations.
    Each image type has a FIXED pixel dimension — no free-form sizes."""
    templates = {
        "poster":   ("professional esports tournament poster {n} {g} neon dark background gold purple cinematic", 1280, 720),
        "roadmap":  ("tournament roadmap timeline {n} {g} stages Registration GroupDraw Schedule MatchDay Champion modern dark", 1280, 720),
        "group":    ("esports group draw reveal {n} {g} {x} dark neon panels", 1024, 576),
        "schedule": ("match schedule card {n} {g} {x} dark professional infographic", 1024, 576),
        "result":   ("match result card {x} {n} dark dramatic victory graphic", 1024, 576),
        "champion": ("champion victory {x} wins {n} {g} golden trophy confetti epic cinematic", 1280, 720),
    }
    tpl, w, h = templates.get(kind, ("professional esports graphic {n} {g}", 1280, 720))
    prompt = tpl.format(n=t_name, g=game, x=extra)
    prompt = prompt.replace(" ", "%20").replace(",", "%2C")
    return (
        "https://image.pollinations.ai/prompt/" + prompt
        + f"?width={w}&height={h}&nologo=true&seed={now_ts()}&model=flux"
    )


#
#  RULES:
#  1. Bot replies to a user ONCE per support session
#  2. Session stays locked until staff marks it resolved in #staff-log
#  3. Uses Pollinations free text AI for natural, accurate responses
#  4. Always mentions real channel names from the server
#  5. Unknown/complex queries → staff-log with full context

import urllib.parse

# ── Per-guild lock: {guild_id: {user_id: support_message_db_id}}
# Prevents double-replies until staff clears the ticket
_replied_users: dict[str, dict[str, str]] = {}

async def ai_generate(prompt: str) -> str:
    """Try Groq (fast) first, fallback to Pollinations (no key needed)."""
    GROQ_KEY = os.environ.get("GROQ_API_KEY", "")

    if GROQ_KEY:
        try:
            async with aiohttp.ClientSession() as s:
                async with s.post("https://api.groq.com/openai/v1/chat/completions",
                    json={"model": "llama-3.3-70b-versatile",
                          "messages": [{"role": "user", "content": prompt}],
                          "max_tokens": 300, "temperature": 0.7},
                    headers={"Authorization": f"Bearer {GROQ_KEY}", "Content-Type": "application/json"},
                    timeout=aiohttp.ClientTimeout(total=12)) as r:
                    if r.status == 200:
                        return (await r.json())["choices"][0]["message"]["content"].strip()
                    print(f"[AI] Groq error {r.status}: {(await r.text())[:200]}", flush=True)
        except Exception as e:
            print(f"[AI] Groq exception: {e}", flush=True)

    try:
        encoded = urllib.parse.quote(prompt[:1500])
        async with aiohttp.ClientSession() as s:
            async with s.get(f"https://text.pollinations.ai/{encoded}",
                timeout=aiohttp.ClientTimeout(total=15)) as r:
                if r.status == 200:
                    return (await r.text()).strip()
                print(f"[AI] Pollinations {r.status}", flush=True)
    except Exception as e:
        print(f"[AI] Pollinations exception: {e}", flush=True)

    return ""

def build_server_context(guild: discord.Guild, active_tournament: dict | None) -> str:
    """Build context string about the server for the AI prompt."""
    channels = ", ".join(f"#{c.name}" for c in guild.text_channels[:30])
    roles = ", ".join(r.name for r in [r for r in guild.roles if not r.is_default()][:20])
    ctx = f"Server: {guild.name}\nChannels: {channels}\nRoles: {roles}\n"
    if active_tournament:
        t = active_tournament
        ctx += (f"\nActive Tournament: {t.get('name','?')}\n"
                f"Game: {t.get('game','?')} | Status: {t.get('status','?')} | "
                f"Date: {t.get('tournament_date','?')} | Time: {t.get('tournament_time','?')}\n"
                f"Slots: {t.get('max_players','?')} | Registered: {t.get('registered_count',0)} | "
                f"Team Size: {t.get('team_size',4)} | Prize: {t.get('prize_pool','?')}\n")
        if t.get('short_name'):
            ctx += f"Registration channel: #{t['short_name']}-register\n"
    else:
        ctx += "\nNo active tournament.\n"
    return ctx

def _xl_header_style(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _xl_border(cell):
    thin = {"border_style": "thin", "color": "AAAAAA"}
    from openpyxl.styles import Border, Side
    cell.border = Border(left=Side(**thin), right=Side(**thin), top=Side(**thin), bottom=Side(**thin))

def _xl_sheet(wb, title, headers, rows, color="1F4E79", widths=None, status_col=0, status_colors=None):
    """Create a styled Excel sheet with headers and data rows."""
    ws = wb.create_sheet(title)
    for ci, h in enumerate(headers, 1):
        _xl_header_style(ws_t := ws.cell(row=1, column=ci, value=h), color)
        _xl_border(ws_t)
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            _xl_border(c)
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 2 and len(headers) <= 4))
    if status_col and status_colors:
        for ri, row in enumerate(rows, 2):
            ws.cell(row=ri, column=status_col).fill = PatternFill("solid", fgColor=status_colors.get(row[status_col-1], "FFFFFF"))
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    return ws

async def build_daily_excel(guild: discord.Guild) -> io.BytesIO:
    """Build a daily Excel report for one guild with 5 sheets."""
    wb = openpyxl.Workbook()
    gid = str(guild.id)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    tournaments = await b44_list("Tournament", {"guild_id": gid})
    all_regs = await b44_list("Registration", {"guild_id": gid})
    support_msgs = await b44_list("SupportMessage", {"guild_id": gid})
    matches = await b44_list("Match", {"guild_id": gid})
    t_map = {t["id"]: t.get("name","?") for t in tournaments}

    # Sheet 1: Tournaments
    ws_t = wb.active
    ws_t.title = "Tournaments"
    t_headers = ["ID","Name","Game","Status","Slots","Registered","Prize","Date","Time","Format","Created"]
    for ci, h in enumerate(t_headers, 1):
        _xl_header_style(ws_t.cell(row=1, column=ci, value=h))
        _xl_border(ws_t.cell(row=1, column=ci))
    t_sc = {"registration_open":"C6EFCE","registration_closed":"FFEB9C","groups_generated":"BDD7EE","scheduled":"FCE4D6","completed":"E2EFDA","deleted":"DDDDDD"}
    for ri, t in enumerate(tournaments, 2):
        row = [t.get("id","")[:8], t.get("name",""), t.get("game",""), t.get("status",""),
               t.get("max_players",""), t.get("registered_count",0), t.get("prize_pool",""),
               t.get("tournament_date",""), t.get("tournament_time",""), t.get("format",""), str(t.get("created_date",""))[:10]]
        for ci, val in enumerate(row, 1):
            _xl_border(ws_t.cell(row=ri, column=ci, value=str(val)))
        ws_t.cell(row=ri, column=4).fill = PatternFill("solid", fgColor=t_sc.get(t.get("status",""),"FFFFFF"))
    for ci, w in enumerate([8,25,14,18,6,10,12,12,12,16,12], 1):
        ws_t.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2-4 via _xl_sheet
    _xl_sheet(wb, "Registrations", ["Team Name","Tournament","Players","Discord User","Reg At","Group","Status","Slot"],
        [[r.get("player_name",""), t_map.get(r.get("tournament_id",""),"?"), r.get("player_username",""),
          r.get("player_discord_tag",""), str(r.get("registered_at",""))[:16], r.get("group_label","—"),
          r.get("status","registered"), r.get("seed_number","")] for r in all_regs],
        "375623", widths=[18,22,30,18,16,8,12,7])

    _xl_sheet(wb, "Support Log", ["ID","Message","Status","Created"],
        [[sm.get("id","")[:8], sm.get("message","")[:200], sm.get("status",""), str(sm.get("created_date",""))[:16]] for sm in support_msgs],
        "843C0C", widths=[8,60,12,16], status_col=3,
        status_colors={"resolved":"C6EFCE","pending":"FFEB9C","escalated":"FCE4D6","dismissed":"DDDDDD","saved":"BDD7EE"})

    _xl_sheet(wb, "Match Results", ["Tournament","Round","P1","P2","Winner","Score","Status"],
        [[t_map.get(m.get("tournament_id",""),"?"), m.get("round_number",""), m.get("player1_username",""),
          m.get("player2_username",""), m.get("winner_username",""),
          f"{m.get('player1_score','')}-{m.get('player2_score','')}" if m.get("player1_score") else "N/A",
          m.get("status","")] for m in matches],
        "4472C4", widths=[22,7,18,18,18,8,12])

    # Sheet 5: Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"NexPlay Daily Report — {guild.name}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = f"Generated: {today} (12:00 AM NPT)"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A1:D1"), ws.merge_cells("A2:D2")
    stats = [("",""),("📊 SUMMARY",""),("Total Tournaments",len(tournaments)),
             ("Active Tournaments",len([t for t in tournaments if t.get("status") not in ("completed","deleted")])),
             ("Total Registrations",len(all_regs)),("Support Messages",len(support_msgs)),
             ("Pending Support",len([s for s in support_msgs if s.get("status")=="pending"])),
             ("Resolved Support",len([s for s in support_msgs if s.get("status")=="resolved"])),
             ("Matches Played",len(matches))]
    for ri, (k, v) in enumerate(stats, 3):
        ws.cell(row=ri, column=1, value=k).font = Font(bold=k.startswith("📊"), size=11)
        ws.cell(row=ri, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def send_daily_logs(bot_instance: "NexPlayBot"):
    """Send daily Excel report to server owners."""
    print("[NexPlay] Sending daily logs...", flush=True)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for srv in await b44_list("Server"):
        gid, owner_id = srv.get("guild_id",""), srv.get("owner_id","")
        guild = bot_instance.get_guild(int(gid)) if gid else None
        if not guild or not owner_id:
            continue
        try:
            xl_buf = await build_daily_excel(guild)
            owner = await bot_instance.fetch_user(int(owner_id))
            dm = await owner.create_dm()
            e = discord.Embed(title=f"📊 NexPlay Daily Report — {guild.name}",
                description=f"Daily report for **{today}** — tournaments, registrations, matches & support.",
                color=0x1F4E79, timestamp=datetime.now(timezone.utc))
            e.set_footer(text="NexPlay Tournament System")
            await dm.send(embed=e, file=discord.File(xl_buf, filename=f"NexPlay_DailyLog_{guild.name}_{today}.xlsx"))
            print(f"[NexPlay] Daily log sent to {owner} for {guild.name}", flush=True)
        except Exception as e:
            print(f"[NexPlay] Daily log error for {gid}: {e}", flush=True)


class StaffLogView(discord.ui.View):

    def __init__(self, uid: str, rec_id: str, guild_locks: dict,
                 user: discord.Member, user_channel: discord.TextChannel):
        super().__init__(timeout=None)
        self.uid, self.rec_id, self.guild_locks = uid, rec_id, guild_locks
        self.user, self.user_channel = user, user_channel
        self.message, self._status = None, "pending"

    async def interaction_check(self, interaction: discord.Interaction) -> bool:
        if not is_staff(interaction.user):
            await interaction.response.send_message("❌ Staff only.", ephemeral=True)
            return False
        return True

    # ── ✅ RESOLVE ────────────────────────────────────────────────────────
    @discord.ui.button(label="✅ Resolve", style=discord.ButtonStyle.success, custom_id="staff_resolve")
    async def btn_resolve(self, interaction: discord.Interaction, button: discord.ui.Button):
        self._set_status(interaction, "resolved", "✅ RESOLVED", 0x00FF00, True)
        try:
            await self.user_channel.send(embed=discord.Embed(
                description=f"Hey {self.user.mention}, your query was resolved! Feel free to ask again. 😊",
                color=0x00FF00))
        except: pass
        await interaction.response.send_message("✅ Resolved & lock cleared.", ephemeral=True)

    # ── 💾 SAVE / UNSAVE TOGGLE ───────────────────────────────────────────
    @discord.ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, custom_id="staff_save")
    async def btn_save(self, interaction: discord.Interaction, button: discord.ui.Button):
        if self._status == "saved":
            self._status = "pending"
            if self.rec_id: await b44_update("SupportMessage", self.rec_id, {"status": "pending"})
            button.label, button.style = "💾 Save", discord.ButtonStyle.secondary
            await interaction.response.edit_message(view=self)
            await interaction.followup.send("📌 Log unsaved.", ephemeral=True)
        else:
            self._status = "saved"
            if self.rec_id: await b44_update("SupportMessage", self.rec_id, {"status": "saved"})
            button.label, button.style = "📌 Saved", discord.ButtonStyle.primary
            await interaction.response.edit_message(view=self)
            await interaction.followup.send("💾 Log saved.", ephemeral=True)

    # ── 🔔 ESCALATE ───────────────────────────────────────────────────────
    @discord.ui.button(label="🔔 Escalate", style=discord.ButtonStyle.danger, custom_id="staff_escalate")
    async def btn_escalate(self, interaction: discord.Interaction, button: discord.ui.Button):
        self._set_status(interaction, "escalated", "🚨 ESCALATED", 0xFF6600, False)
        await interaction.response.send_message(f"🚨 Escalated! <@&{interaction.guild.id}> please review.", ephemeral=True)

    # ── 🗑️ DISMISS ────────────────────────────────────────────────────────
    @discord.ui.button(label="🗑️ Dismiss", style=discord.ButtonStyle.secondary, custom_id="staff_dismiss")
    async def btn_dismiss(self, interaction: discord.Interaction, button: discord.ui.Button):
        self._set_status(interaction, "dismissed", "🗑️ DISMISSED", 0x555555, True)
        self.stop()
        await interaction.response.send_message("🗑️ Dismissed & lock cleared.", ephemeral=True)

    def _set_status(self, interaction, status, title, color, clear_lock):
        """Update DB status, rebuild embed with disabled buttons, optionally clear user lock."""
        self._status = status
        if clear_lock:
            self.guild_locks.pop(self.uid, None)
        if self.rec_id:
            asyncio.ensure_future(b44_update("SupportMessage", self.rec_id, {"status": status}))
        if not self.message or not self.message.embeds:
            return
        old = self.message.embeds[0]
        new_e = discord.Embed(title=title, description=old.description, color=color, timestamp=datetime.now(timezone.utc))
        for f in old.fields:
            new_e.add_field(name=f.name, value=f.value, inline=f.inline)
        new_e.set_footer(text=f"By {interaction.user.display_name}")
        for child in self.children:
            child.disabled = True
        try: asyncio.ensure_future(self.message.edit(embed=new_e, view=self))
        except: pass


async def _post_staff_log(message, uid, q, rec_id, guild_locks, needs_staff, ai_reply):
    ch = resolve_channel(message.guild, 'staff')
    if not ch:
        log(f"[Support] No staff-log channel in {message.guild.name}")
        return

    e = discord.Embed(
        title='🔴 NEEDS STAFF REVIEW' if needs_staff else '🟢 AI HANDLED — Monitor',
        color=0xFF4444 if needs_staff else 0x00CC66)
    e.add_field(name='👤 User', value=message.author.mention, inline=True)
    e.add_field(name='📢 Channel', value=message.channel.mention, inline=True)
    e.add_field(name='❓ Message', value=f'```{q[:500]}```', inline=False)
    if needs_staff:
        reason = ai_reply.replace("NEEDS_STAFF:", "").strip() if ai_reply else "No AI response"
        e.add_field(name='⚠️ Reason', value=reason or "Complex query", inline=False)
        e.add_field(name='📋 Action', value='1. Reply to user\n2. React ✅ to clear lock\n3. User can ask again', inline=False)
    else:
        e.add_field(name='🤖 AI Response', value=f'```{ai_reply[:300]}```', inline=False)
        e.add_field(name='📋 Note', value='AI handled. React ✅ to clear lock.', inline=False)
    e.set_footer(text=f"User ID: {uid} | DB Record: {rec_id}")

    try:
        view = StaffLogView(uid=uid, rec_id=rec_id, guild_locks=guild_locks,
                           user=message.author, user_channel=message.channel)
        log_msg = await ch.send(embed=e, view=view)
        view.message = log_msg
    except Exception as ex:
        print(f"[Support] staff-log error: {ex}", flush=True)

    if needs_staff:
        try: await ch.send(f"@here 🔴 **{message.author.display_name}** needs help in {message.channel.mention}")
        except: pass


async def handle_support(message: discord.Message) -> None:
    ok, reason = await check_feature(str(message.guild.id), "ai_support")
    if not ok:
        srv = await get_server_record(str(message.guild.id))
        if srv:
            desc = f"👋 AI support is **Elite** only. You're on **{srv.get('plan_name','?')}** — a staff member will help shortly."
        else:
            desc = "👋 I'm NexPlay Bot. This server isn't registered — ask an admin to run `/setup`."
        await message.reply(embed=discord.Embed(description=desc, color=0xFFA500).set_footer(text="NexPlay Support"), mention_author=False)
        return
    gid = str(message.guild.id)
    uid = str(message.author.id)
    q   = message.content.strip()

    # ── ONE-REPLY LOCK ─────────────────────────────────────────────────────
    guild_locks = _replied_users.setdefault(gid, {})
    if uid in guild_locks:
        # Already replied — silently ignore until staff clears
        return

    # ── Fetch active tournament ────────────────────────────────────────────
    all_t  = await b44_list('Tournament', {'guild_id': gid})
    active = next(
        (t for t in sorted(all_t, key=lambda x: x.get('created_date',''), reverse=True)
         if t.get('status') not in ('completed','cancelled','deleted')),
        None
    )

    # ── Build AI prompt with full server context ───────────────────────────
    server_ctx = build_server_context(message.guild, active)

    prompt = f"""You are NexPlay Bot, a friendly and helpful Discord assistant for a gaming tournament server.

SERVER CONTEXT:
{server_ctx}

USER MESSAGE from {message.author.display_name}:
"{q}"

INSTRUCTIONS:
- Reply in a friendly, warm, and accurate way (2-4 sentences max)
- If they ask about a channel, mention the EXACT channel name from the server context above (e.g. #ff-register)
- If they ask about registration, tell them which channel to go to and the format
- If they ask about schedule/date/time, give the exact info from tournament data
- If they ask about rules, summarize the key rules
- If you don't know the answer or it needs staff action, say: "NEEDS_STAFF: <brief reason>"
- NEVER make up information. Only use data from the context above.
- Keep it short, casual, helpful. Use emojis sparingly."""

    ai_reply = await ai_generate(prompt)

    needs_staff = not ai_reply or ai_reply.startswith("NEEDS_STAFF") or len(ai_reply) < 10

    ql = q.lower()
    SUPPORT_COLORS = [(0x00FF7F,['register','join','sign up']),(0x1E90FF,['when','time','date','schedule']),
                      (0xFFD700,['prize','reward','money','win']),(0xFF4444,['rule','cheat','banned','unfair']),
                      (0x9B59B6,['bracket','group','opponent'])]
    color = next((c for c, kws in SUPPORT_COLORS if any(k in ql for k in kws)), 0x5865F2)

    # ── Reply to user ──────────────────────────────────────────────────────
    if not needs_staff and ai_reply:
        clean_reply = ai_reply.replace("NEEDS_STAFF:", "").strip()
        embed = discord.Embed(description=clean_reply, color=color)
        embed.set_footer(text="NexPlay Support • Reply handled by AI")
        await message.reply(embed=embed, mention_author=False)

    # ── Save to DB ────────────────────────────────────────────────────────
    support_rec = await b44_create('SupportMessage', {
        'guild_id':   gid,
        'guild_name': message.guild.name,
        'message':    q,
        'status':     'pending' if needs_staff else 'ai_resolved',
    })
    rec_id = support_rec.get('id', '')

    # ── Lock this user from getting another reply ─────────────────────────
    guild_locks[uid] = rec_id

    await _post_staff_log(message, uid, q, rec_id, guild_locks, needs_staff, ai_reply)
async def find_tournament_by_short(guild_id: str, short_name: str, statuses: tuple = ("registration_open",)) -> dict | None:
    """Find a tournament by short_name in a guild."""
    for t in await b44_list("Tournament", {"guild_id": guild_id}):
        if t.get("short_name", "").lower() == short_name.lower() and t.get("status") in statuses:
            return t
    return None

async def _find_tournament(guild_id: str, name: str) -> dict | None:
    """Find a tournament by name (case-insensitive) or short_name in a guild."""
    for t in await b44_list("Tournament", {"guild_id": guild_id}):
        if t.get("name", "").lower() == name.lower() or t.get("short_name", "").lower() == name.lower():
            return t
    return None


def parse_registration(text: str, team_size: int) -> dict | None:
    """
    Parse a team registration message. Returns dict or None if invalid.
    Expected format (case-insensitive keys):
        Team Name: <name>
        Player 1: @mention
        Player 2: @mention
        ...
    """
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    result = {}

    # Find team name line
    team_line = next((l for l in lines if re.match(r"team\s*name\s*:", l, re.I)), None)
    if not team_line:
        return None
    result["team_name"] = re.sub(r"^team\s*name\s*:\s*", "", team_line, flags=re.I).strip()
    if not result["team_name"]:
        return None

    # Find player lines
    players = []
    for i in range(1, team_size + 1):
        pat = re.compile(r"player\s*" + str(i) + r"\s*:\s*(.*)", re.I)
        pline = next((l for l in lines if pat.match(l)), None)
        if not pline:
            return None
        raw = pat.match(pline).group(1).strip()
        # Extract mention user ID
        uid_match = re.search(r"<@!?(\d+)>", raw)
        if not uid_match:
            return None
        players.append(uid_match.group(1))

    if len(players) != team_size:
        return None

    result["players"] = players
    return result


async def lock_register_channel(guild: discord.Guild, channel: discord.TextChannel):
    try:
        everyone = guild.default_role
        overwrite = channel.overwrites_for(everyone)
        overwrite.send_messages = False
        await channel.set_permissions(everyone, overwrite=overwrite, reason="NexPlay: Registration full")
    except Exception as e:
        log(f"[WARN] Could not lock channel #{channel.name}: {e}")


async def update_reg_announcement(tournament: dict, guild: discord.Guild, registered: int):
    msg_id = tournament.get("registration_msg_id")
    ch_id  = tournament.get("registration_channel_id")
    if not msg_id or not ch_id:
        return
    try:
        ch = guild.get_channel(int(ch_id))
        if not ch:
            return
        msg = await ch.fetch_message(int(msg_id))
        if not msg:
            return
        # Rebuild embed with updated count
        t = tournament
        max_p = t.get("max_players", 16)
        name  = t.get("name", "Tournament")
        game  = t.get("game", "")
        prize = t.get("prize_pool", "")
        date  = t.get("tournament_date", "")
        tsize = t.get("team_size", 4)
        short = t.get("short_name", "t")
        bar   = "█" * registered + "░" * (max_p - registered)
        filled = registered >= max_p

        lines = "\n".join([f"Player {i+1}: @mention" for i in range(tsize)])
        if filled:
            desc = (f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n🎮 **Game:** {game}\n🎖️ **Prize:** {prize}\n"
                    f"📅 **Date:** {date}\n👥 **Slots:** {registered}/{max_p}  `{bar}`\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n🚫 **Registration CLOSED. All slots filled!**")
        else:
            desc = (f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n🎮 **Game:** {game}\n🎖️ **Prize:** {prize}\n"
                    f"📅 **Date:** {date}\n👥 **Slots:** {registered}/{max_p}  `{bar}`\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n@everyone **Registration OPEN!**\n"
                    f"```\nTeam Name: <name>\n{lines}\n```\n⚠️ All players must be @mentioned.")
        e = discord.Embed(title=f"{'🔒 CLOSED' if filled else '✍️ OPEN'} — {name}", description=desc,
            color=0x555555 if filled else 0x00FF7F, timestamp=datetime.now(timezone.utc))
        e.set_footer(text="NexPlay")
        await msg.edit(embed=e)
    except Exception as e:
        log(f"[WARN] Could not update reg announcement: {e}")

async def handle_registration(message: discord.Message, gid: str, short: str, tournament: dict):
    text = message.content.strip()
    team_size = int(tournament.get("team_size", 4))
    parsed = parse_registration(text, team_size)

    if not parsed:
        lines_needed = "\n".join([f"Player {i+1}: @mention" for i in range(team_size)])
        await _reject_msg(message, "❌ Invalid Format",
            f"Use EXACTLY this format:\n```\nTeam Name: <name>\n{lines_needed}\n```"
            f"\n• All {team_size} players must be @mentioned")
        return

    team_name = parsed["team_name"]
    players = parsed["players"]

    existing_regs = await b44_list("Registration", {"tournament_id": tournament["id"]})

    # Duplicate team name?
    if any(r.get("player_name", "").lower() == team_name.lower() for r in existing_regs):
        await _reject_msg(message, "❌ Team Name Taken", f"**{team_name}** is already registered.")
        return

    # Duplicate player?
    all_players = []
    for r in existing_regs:
        members = r.get("team_members", [])
        if isinstance(members, list): all_players.extend(members)
        elif isinstance(members, str) and members: all_players.extend(members.split(","))
    already = [f"<@{p}>" for p in players if p in all_players]
    if already:
        await _reject_msg(message, "❌ Player Already Registered", f"{', '.join(already)} is already on another team!")
        return

    # Slots full?
    max_p = int(tournament.get("max_players", 16))
    slot = len(existing_regs) + 1
    if slot > max_p:
        await message.add_reaction("❌")
        await message.reply(embed=discord.Embed(title="🔒 Registration Full",
            description=f"All {max_p} slots are taken.", color=0xFF4444))
        return

    # Save registration
    await b44_create("Registration", {
        "tournament_id": tournament["id"], "guild_id": gid,
        "player_name": team_name, "player_discord_id": str(message.author.id),
        "team_members": players, "status": "registered", "logo_url": "",
    })
    await b44_update("Tournament", tournament["id"], {"registered_count": slot})
    await message.add_reaction("✅")

    # Post confirmation to #<short>-confirm-teams
    cfm_ch = discord.utils.get(message.guild.text_channels, name=short + "-confirm-teams")
    if cfm_ch:
        player_mentions = " ".join(f"<@{p}>" for p in players)
        logo_ch = discord.utils.get(message.guild.text_channels, name=short + "-team-logo")
        logo_mention = f"<#{logo_ch.id}>" if logo_ch else f"#{short}-team-logo"
        e = discord.Embed(title="✅ TEAM REGISTERED!",
            description=(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"🏷️ **Team:** {team_name}\n👑 **Captain:** {message.author.mention}\n"
                f"👥 **Players:** {player_mentions}\n🎫 **Slot:** #{slot} of {max_p}\n"
                f"🏆 **Tournament:** {tournament.get('name', '')}\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n**CONFIRMED ✅**\n\n"
                f"🎨 Submit logo in {logo_mention} — `Team Name: <name>` + image"),
            color=0x00FF7F, timestamp=datetime.now(timezone.utc))
        e.set_footer(text=f"NexPlay | {tournament.get('game','')}", icon_url="https://i.imgur.com/wSTFkRM.png")
        await cfm_ch.send(embed=e)

    # Update announcement + auto-close
    updated_t = dict(tournament); updated_t["registered_count"] = slot
    await update_reg_announcement(updated_t, message.guild, slot)
    if slot >= max_p:
        await b44_update("Tournament", tournament["id"], {"status": "registration_closed"})
        await lock_register_channel(message.guild, message.channel)
        await message.channel.send(embed=discord.Embed(
            title=f"🔒 REGISTRATION CLOSED — {tournament.get('name', '')}",
            description=f"All **{max_p}** slots filled! Groups next. Stay tuned.",
            color=0xFF6B35, timestamp=datetime.now(timezone.utc)).set_footer(text="NexPlay"))


async def handle_logo_submission(message: discord.Message, gid: str, short: str, tournament: dict):
    images = [a for a in message.attachments if a.content_type and a.content_type.startswith("image/")]
    if not images:
        await _reject_msg(message, "❌ No Image Attached",
            "Attach your team logo (PNG/JPG) + include your team name.\n"
            "```\nTeam Name: <exact team name>\n[attach image]\n```")
        return

    match = re.search(r'team\s*name\s*:\s*(.+)', message.content, re.IGNORECASE)
    if not match:
        await _reject_msg(message, "❌ Missing Team Name", "Include `Team Name: <name>` in your message.")
        return

    submitted_name = match.group(1).strip()
    existing_regs = await b44_list("Registration", {"tournament_id": tournament["id"]})
    matched = next((r for r in existing_regs if r.get("player_name", "").lower() == submitted_name.lower()), None)
    if not matched:
        matched = next((r for r in existing_regs if r.get("player_discord_id") == str(message.author.id)), None)
        if matched: submitted_name = matched.get("player_name", submitted_name)
    if not matched:
        await _reject_msg(message, "❌ Team Not Found",
            f"**{submitted_name}** is not registered in this tournament.", delay=20)
        return

    logo_url = images[0].url
    await b44_update("Registration", matched["id"], {
        "logo_url": logo_url, "logo_submitted_at": datetime.now(timezone.utc).isoformat(),
        "logo_submitted_by": str(message.author.id),
    })
    await message.add_reaction("✅")

    e = discord.Embed(title="✅ Logo Accepted!",
        description=f"**Team:** {submitted_name}\n**By:** {message.author.mention}\n"
            f"Saved — will appear in group draws and match cards.",
        color=0x00FF7F, timestamp=datetime.now(timezone.utc))
    e.set_thumbnail(url=logo_url)
    e.set_footer(text="NexPlay")
    await message.channel.send(embed=e)

    cfm_ch = discord.utils.get(message.guild.text_channels, name=short + "-confirm-teams")
    if cfm_ch:
        e2 = discord.Embed(title="🎨 Logo Submitted",
            description=f"**{submitted_name}** submitted their logo ✅\nBy: {message.author.mention}",
            color=0xFF6B9D, timestamp=datetime.now(timezone.utc))
        e2.set_thumbnail(url=logo_url)
        e2.set_footer(text="NexPlay | Logo will appear in all GFX")
        await cfm_ch.send(embed=e2)


@bot.event
async def on_message(message: discord.Message):
    if message.author.bot:
        return

    if not message.guild:
        await bot.process_commands(message)
        return

    # ── Debug log every message so we can verify receipt ─────────────────────
    print(f"[MSG] #{message.channel.name} | {message.author} | {message.content[:80]}", flush=True)

    # ── Support handler — ONLY in designated support/help channels ──────────────
    ch_name_lower = message.channel.name.lower()

    SUPPORT_NAMES = (
        "support", "support-ticket", "help", "nexplay-support",
        "nexplay-help", "bot-help", "ask-here", "questions",
    )
    is_support_ch = (
        ch_name_lower in SUPPORT_NAMES
        or ch_name_lower.startswith("support")
        or ch_name_lower.startswith("help")
        or ch_name_lower.startswith("ask")
    )

    if is_support_ch and isinstance(message.channel, discord.TextChannel) and len(message.content.strip()) > 3:
        await handle_support(message)
        # Don't return — let process_commands run too, so slash commands still work

    # ── Registration channel handler ──────────────────────────────────────────
    gid = str(message.guild.id)
    ch_name = message.channel.name  # e.g. "npo26-register"

    if ch_name.endswith("-register"):
        short_candidate = ch_name[:-9]
        tournament = await find_tournament_by_short(gid, short_candidate, ("registration_open",))
        if tournament:
            await handle_registration(message, gid, short_candidate, tournament)
            return
        await bot.process_commands(message)
        return

    if ch_name.endswith("-team-logo"):
        short_candidate = ch_name[:-10]
        tournament = await find_tournament_by_short(gid, short_candidate, ("registration_open", "registration_closed"))
        if tournament:
            await handle_logo_submission(message, gid, short_candidate, tournament)
            return
        await bot.process_commands(message)
        return

    await bot.process_commands(message)


# ── /setup ────────────────────────────────────────────────
@tree.command(name="clearlog", description="[Staff] Clear a user support lock so they can ask again")
@app_commands.describe(user="The user whose support lock to clear")
async def clearlog(interaction: discord.Interaction, user: discord.Member):
    if not is_staff(interaction.user):
        return await interaction.response.send_message("❌ Staff only.", ephemeral=True)
    gid, uid = str(interaction.guild.id), str(user.id)
    guild_locks = _replied_users.get(gid, {})
    if uid not in guild_locks:
        return await interaction.response.send_message(f"✅ {user.mention} has no active lock.", ephemeral=True)
    rec_id = guild_locks.pop(uid, None)
    if rec_id:
        try: await b44_update('SupportMessage', rec_id, {'status': 'resolved'})
        except: pass
    await interaction.response.send_message(embed=ok_e("🔓 Lock Cleared", f"{user.mention} can now ask again."), ephemeral=True)
    try: await interaction.channel.send(embed=discord.Embed(description=f"Hey {user.mention}, your support session was cleared! 😊", color=0x00FF00))
    except: pass


@tree.command(name="setup", description="Initialize NexPlay in this server (server owner only)")
async def cmd_setup(interaction: discord.Interaction):
    if interaction.user.id != interaction.guild.owner_id and not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("Only the server owner can run /setup."), ephemeral=True)
    await interaction.response.defer(thinking=True, ephemeral=True)
    gid = str(interaction.guild.id)

    # Register or update server record
    existing = await get_server_record(gid)
    if not existing:
        # No record at all — try to create one
        if not SVC_TOKEN:
            status_msg = "⚠️ **Database not configured.** `BASE44_SERVICE_TOKEN` env var is missing on the bot server. Contact the bot owner."
        else:
            created = await register_server(interaction.guild)
            if created and created.get("id"):
                status_msg = "Server registered! You have a **free trial** with 3 tournaments."
            else:
                status_msg = "⚠️ Failed to register server in the database. Check bot logs for the error."
    elif existing.get("subscription_status") == "inactive":
        # Server was previously removed — reactivate, preserving plan + tournament history
        await b44_update("Server", existing["id"], {
            "subscription_status": "trial" if existing.get("plan_name", "Free Trial") == "Free Trial" else "active",
            "guild_name": interaction.guild.name,
            "owner_id": str(interaction.guild.owner.id) if interaction.guild.owner else existing.get("owner_id", ""),
            "owner_name": interaction.guild.owner.display_name if interaction.guild.owner else existing.get("owner_name", ""),
            "member_count": interaction.guild.member_count or 0,
            "last_active": now_iso(),
        })
        plan = existing.get("plan_name", "Free Trial")
        t_used = existing.get("tournaments_used", 0)
        status_msg = f"♻️ **Welcome back!** Your server has been reactivated.\nPlan: **{plan}** | Tournaments used: **{t_used}**\nAll your previous data has been preserved."
    else:
        status_msg = "Server already registered. Status: **" + str(existing.get("subscription_status", "?")) + "**"

    # Provision channels
    channels = await get_or_create_channels(interaction.guild)
    ch_list  = "\n".join(f"> <#{v}>" for v in channels.values() if v) or "No channels created (check bot permissions)."

    await interaction.followup.send(embed=ok_e(
        "NexPlay Setup Complete!",
        status_msg + "\n\n**Channels ready:**\n" + ch_list + "\n\n"
        "Use `/create_tournament` to launch your first tournament!\n"
        "\n**Links:**\n"
        "Portal: https://nexplay-server-portal.vercel.app\n"
        "Billing: https://nexplay-server-portal.vercel.app/subscription"
    ), ephemeral=True)


# ── /create_tournament ────────────────────────────────────

async def make_tournament_channels(guild: discord.Guild, tournament_name: str) -> dict:
    """Create a tournament category + 8 channels. Returns dict with channel IDs."""
    
    # Generate short name: initials + numbers, max 6 chars
    words = tournament_name.strip().split()
    if len(words) == 1:
        short = words[0][:4].lower()
    else:
        short = ("".join(w[0] for w in words) + re.sub(r"[^0-9]", "", tournament_name)[-2:])[:6].lower()
    short = re.sub(r"[^a-z0-9]", "", short)[:6] or "tourney"

    cat_name = f"🏆 {short.upper()}"
    
    # Sub-channels: (key, channel_suffix, topic)
    channels_def = [
        ("info",          f"{short}-info",           "📋 Tournament information, rules and details"),
        ("announcements", f"{short}-announcements",   "📢 Official tournament announcements"),
        ("roadmap",       f"{short}-roadmap",         "🗺️ Tournament roadmap and schedule overview"),
        ("results",       f"{short}-results",         "🏅 Match results and standings"),
        ("groups",        f"{short}-groups",          "🎯 Group draws and bracket reveal"),
        ("register",      f"{short}-register",        "✍️ Player registration — type Team Name: + Player 1: @mention ..."),
        ("confirm-teams", f"{short}-confirm-teams",   "✅ Team confirmation and roster lock"),
        ("team-logo",     f"{short}-team-logo",       "🎨 Submit team logo — type Team Name: <name> + attach image"),
        ("help",          f"{short}-help",            "❓ Support and help for this tournament"),
    ]
    
    result = {"short_name": short, "category_name": cat_name}
    
    # ── Create or find category ───────────────────────────────────────────────
    category = discord.utils.get(guild.categories, name=cat_name)
    if not category:
        try:
            category = await guild.create_category(
                cat_name,
                reason=f"NexPlay: Tournament '{tournament_name}' channels"
            )
        except Exception as e:
            log(f"[ERROR] Could not create category {cat_name}: {e}")
            category = None
    
    result["category_id"] = str(category.id) if category else None

    # ── Create each channel ───────────────────────────────────────────────────
    for key, ch_name, topic in channels_def:
        existing = discord.utils.get(guild.text_channels, name=ch_name)
        if not existing:
            try:
                ch = await guild.create_text_channel(
                    ch_name,
                    category=category,
                    topic=topic,
                    reason=f"NexPlay: {tournament_name} tournament"
                )
                result[key] = str(ch.id)
            except Exception as e:
                log(f"[ERROR] Could not create #{ch_name}: {e}")
                result[key] = None
        else:
            result[key] = str(existing.id)
    
    log(f"[Channels] Created tournament channels for '{tournament_name}' → category '{cat_name}'")
    return result


@tree.command(name="create_tournament", description="[Host] Create a tournament — opens a 3-step form")
async def cmd_create(interaction: discord.Interaction):
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    allowed, reason = await is_allowed(str(interaction.guild.id), interaction.guild)
    if not allowed:
        return await interaction.response.send_message(embed=err_e(reason), ephemeral=True)
    # Open step 1 modal
    await interaction.response.send_modal(TournamentStep1Modal())


@tree.command(name="edit_tournament", description="[Host] Edit an existing tournament details")
@app_commands.describe(tournament_name="Name of the tournament to edit")
async def cmd_edit_tournament_new(interaction: discord.Interaction, tournament_name: str):
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)

    # Feature gate — Starter+
    await interaction.response.defer(thinking=True, ephemeral=True)
    ok, msg = await check_feature(str(interaction.guild.id), "edit_tournament")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    tournament = await _find_tournament(gid, tournament_name)
    if not tournament:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament_name}** not found."), ephemeral=True)

    view = TournamentEditSlotView(tournament)
    info_embed = discord.Embed(
        title=f"✏️ Edit: {tournament['name']}",
        description=(
            f"**Current values:**\n"
            f"🎮 Game: {tournament.get('game','')}\n"
            f"💰 Prize: {tournament.get('prize_pool','')}\n"
            f"📅 Date: {tournament.get('tournament_date','')}  |  ⏰ {tournament.get('tournament_time','')}\n"
            f"👥 Max Teams: {tournament.get('max_players','')}  |  Team Size: {tournament.get('team_size','')}v{tournament.get('team_size','')}\n"
            f"🎯 Group Size: {tournament.get('group_size','')}  |  Rounds: {tournament.get('rounds','')}\n"
            f"🌏 Nations: {tournament.get('eligible_nations','🇳🇵')}\n"
            f"📊 Status: {tournament.get('status','')}\n\n"
            f"Choose what to edit below:"
        ),
        color=0xFFA500
    )
    await interaction.followup.send(embed=info_embed, view=view, ephemeral=True)


# ────────────────────────────────────────────────────────────
#  ELITE FEATURE COMMANDS
# ────────────────────────────────────────────────────────────

@tree.command(name="host_game", description="[Elite] Host a mini-game event (Trivia, Prediction, GG Hunt)")
@app_commands.describe(game_type="Type of mini-game to host")
@app_commands.choices(game_type=[
    app_commands.Choice(name="🧠 Trivia Challenge",      value="trivia"),
    app_commands.Choice(name="🎯 Prediction Game",       value="prediction"),
    app_commands.Choice(name="🔍 GG Hunt (hidden word)", value="gg_hunt"),
])
async def cmd_host_game(interaction: discord.Interaction, game_type: str):
    if not await _staff_feature_gate(interaction, "host_game"):
        return

    games = {
        "trivia":     ("🧠 TRIVIA CHALLENGE", "What is the max players in a Free Fire match?", "50", "It's between 40 and 60.", 0x3498DB),
        "prediction":  ("🎯 PREDICTION GAME", "Who will win the next NexPlay tournament?", None, "No wrong answers — just fun!", 0xE74C3C),
        "gg_hunt":    ("🔍 GG HUNT", "Find the hidden word: 'Great players always GRIND for the Glory!'", "GRIND", "It's in ALL CAPS.", 0x2ECC71),
    }
    title, q, a, hint, color = games.get(game_type, games["trivia"])
    desc = f"**❓ {q}**\n\n💡 Hint: *{hint}*\n\n" + ("Type your answer! First correct wins 🏆" if a else "Drop your prediction! 👇")
    e = discord.Embed(title=title, description=desc, color=color, timestamp=datetime.now(timezone.utc))
    e.set_footer(text=f"NexPlay Mini-Game | {interaction.user.display_name}")
    await interaction.followup.send("@everyone", embed=e)


@tree.command(name="suggest_improvement", description="[Elite] Get AI-powered server growth suggestions")
async def cmd_suggest_improvement(interaction: discord.Interaction):
    if not await _staff_feature_gate(interaction, "suggest_improvement"):
        return

    gid = str(interaction.guild.id)
    guild = interaction.guild
    active_t = None
    all_ts = await b44_list("Tournament", {"guild_id": gid})
    if all_ts:
        active_t = next((t for t in all_ts if t.get("status") in ("registration_open","in_progress")), all_ts[0])

    ctx = build_server_context(guild, active_t)
    prompt = (f"You are a Discord server growth expert for NexPlay esports.\n"
              f"Server context:\n{ctx}\n\n"
              f"Give 5 actionable suggestions to grow this server and increase tournament engagement. "
              f"Be specific and enthusiastic. Numbered list.")
    suggestions = await ai_generate(prompt)
    e = discord.Embed(title="💡 NexPlay Growth Advisor", description=suggestions[:4000],
        color=0xF39C12, timestamp=datetime.now(timezone.utc))
    e.set_footer(text="NexPlay AI | Elite Plan")
    await interaction.followup.send(embed=e)


async def build_tournament_export_excel(tournament: dict, registrations: list, groups: list, matches: list) -> io.BytesIO:
    """Build an Excel backup for a single tournament — 4 sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Tournament details
    ws_t = wb.active
    ws_t.title = "Tournament"
    fields = ["id","name","short_name","game","status","format","prize_pool","description","rules",
              "max_players","team_size","group_size","rounds","registered_count","tournament_date",
              "tournament_time","reg_deadline","stream_channel","eligible_nations","category_name",
              "announcement_channel_id","registration_channel_id","created_by_discord_id","created_date","started_at"]
    labels = ["ID","Name","Short Name","Game","Status","Format","Prize Pool","Description","Rules",
              "Max Players","Team Size","Group Size","Rounds","Registered","Date","Time","Reg Deadline",
              "Stream","Nations","Category","Ann Ch ID","Reg Ch ID","Created By","Created At","Started At"]
    for ci, h in enumerate(["Field","Value"], 1):
        _xl_header_style(ws_t.cell(row=1, column=ci, value=h))
    for ri, (label, key) in enumerate(zip(labels, fields), 2):
        ws_t.cell(row=ri, column=1, value=label).font = Font(bold=True)
        _xl_border(ws_t.cell(row=ri, column=1))
        c2 = ws_t.cell(row=ri, column=2, value=str(tournament.get(key, "")))
        _xl_border(c2)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws_t.column_dimensions["A"].width = 24
    ws_t.column_dimensions["B"].width = 50

    # Sheets 2-4 via _xl_sheet
    _xl_sheet(wb, "Registrations", ["Team Name","Discord User","Team Members","Status","Seed #","Group","Logo URL","Registered At"],
        [[r.get("player_name",""), r.get("player_discord_id",""), str(r.get("team_members","")),
          r.get("status",""), r.get("seed_number",""), r.get("group_label","—"),
          r.get("logo_url",""), str(r.get("registered_at",""))] for r in registrations],
        "375623", widths=[18,18,35,14,7,8,30,20])

    _xl_sheet(wb, "Groups", ["Group Label","Player Names","Player IDs","Generated At"],
        [[g.get("group_label",""), str(g.get("player_names","")), str(g.get("player_ids","")), str(g.get("generated_at",""))] for g in groups],
        "843C0C", widths=[12,40,40,20])

    _xl_sheet(wb, "Matches", ["Round","Match #","Group","P1","P2","P1 Score","P2 Score","Winner","Status","Scheduled","Results Card"],
        [[m.get("round_number",""), m.get("match_number",""), m.get("group_label","—"),
          m.get("player1_username",""), m.get("player2_username",""), m.get("player1_score",""),
          m.get("player2_score",""), m.get("winner_username",""), m.get("status",""),
          str(m.get("scheduled_at","")), m.get("results_card_image_url","")] for m in matches],
        "1F4E79", widths=[7,8,8,18,18,8,8,18,12,20,30])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════
#  TOURNAMENT MANAGEMENT COMMANDS
# ═══════════════════════════════════════════════════════════════════════════

@tree.command(name="register", description="✍️ Register your team for a tournament")
@app_commands.describe(tournament="Tournament name (or leave blank if in a #xxx-register channel)")
async def cmd_register(interaction: discord.Interaction, tournament: str = ""):
    """Show registration instructions for a tournament."""
    gid = str(interaction.guild.id)
    # Try to find tournament by name or by current channel
    t = None
    if tournament:
        t = await _find_tournament(gid, tournament)
    else:
        # Detect from channel name: <short>-register
        ch = interaction.channel.name
        if ch.endswith("-register"):
            short = ch[:-9]
            t = await find_tournament_by_short(gid, short, ("registration_open",))
        if not t:
            # List open tournaments
            all_t = await b44_list("Tournament", {"guild_id": gid})
            open_ts = [x for x in all_t if x.get("status") == "registration_open"]
            if len(open_ts) == 1:
                t = open_ts[0]
            elif len(open_ts) > 1:
                names = "\n".join(f"• **{x['name']}** — Register in #{x.get('short_name','')}-register" for x in open_ts[:10])
                return await interaction.response.send_message(embed=err_e(
                    f"Multiple tournaments open. Specify one:\n\n{names}"), ephemeral=True)

    if not t:
        return await interaction.response.send_message(embed=err_e(
            "No open tournament found. Go to a #<short>-register channel or specify a tournament name."), ephemeral=True)

    if t.get("status") != "registration_open":
        return await interaction.response.send_message(embed=err_e(
            f"**{t['name']}** registration is **{t.get('status','closed')}**."), ephemeral=True)

    team_size = int(t.get("team_size", 4))
    max_p = int(t.get("max_players", 16))
    registered = int(t.get("registered_count", 0))
    slots_left = max_p - registered
    lines = "\n".join([f"Player {i+1}: @mention" for i in range(team_size)])
    short = t.get("short_name", "")

    e = discord.Embed(title=f"✍️ Register for {t['name']}", color=0xFFD700, timestamp=datetime.now(timezone.utc))
    e.description = (
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🎮 **Game:** {t.get('game','')}\n"
        f"💰 **Prize:** {t.get('prize_pool','')}\n"
        f"📅 **Date:** {t.get('tournament_date','')}\n"
        f"👥 **Slots:** {registered}/{max_p} ({slots_left} left)\n"
        f"🏷️ **Team Size:** {team_size}v{team_size}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"**📋 FORMAT — Copy & fill in this #register channel:**\n"
        f"```\nTeam Name: <your team name>\n{lines}\n```\n"
        f"⚠️ All {team_size} players MUST be @mentioned.")
    e.set_footer(text="NexPlay Registration")
    await interaction.response.send_message(embed=e)


@tree.command(name="groups", description="🎯 Generate group draws for a tournament")
@app_commands.describe(tournament="Tournament name")
async def cmd_groups(interaction: discord.Interaction, tournament: str):
    """Generate random group draws from registered teams."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), "groups")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    if t.get("status") not in ("registration_closed", "registration_open"):
        return await interaction.followup.send(embed=err_e(f"Close registration first. Current status: **{t.get('status')}**"), ephemeral=True)

    regs = await b44_list("Registration", {"tournament_id": t["id"]})
    if len(regs) < 2:
        return await interaction.followup.send(embed=err_e("Need at least 2 registered teams to generate groups."), ephemeral=True)

    group_size = int(t.get("group_size", 4))
    teams = list(regs)
    random.shuffle(teams)
    num_groups = max(1, (len(teams) + group_size - 1) // group_size)
    groups = [[] for _ in range(num_groups)]
    for i, team in enumerate(teams):
        groups[i % num_groups].append(team)

    # Save to DB + post to #groups channel
    short = t.get("short_name", "")
    groups_ch = discord.utils.get(interaction.guild.text_channels, name=f"{short}-groups")
    desc = ""
    for gi, group in enumerate(groups):
        label = chr(65 + gi)  # A, B, C...
        player_names = [r.get("player_name", "?") for r in group]
        player_ids = [r.get("player_discord_id", "") for r in group]
        desc += f"**🏆 Group {label}**\n" + "\n".join(f"{i+1}. {n}" for i, n in enumerate(player_names)) + "\n\n"
        # Save group to DB
        await b44_create("TournamentGroup", {
            "tournament_id": t["id"], "guild_id": gid,
            "group_label": label,
            "player_ids": player_ids,
            "player_names": player_names,
            "generated_at": now_iso(),
        })
        # Update registration records with group label
        for r in group:
            await b44_update("Registration", r["id"], {"group_label": label, "seed_number": groups.index(group) * group_size + group.index(r) + 1})

    # Update tournament status
    await b44_update("Tournament", t["id"], {"status": "groups_generated"})

    e = discord.Embed(title=f"🎯 Group Draw — {t['name']}", description=desc, color=0x5865F2, timestamp=datetime.now(timezone.utc))
    e.set_footer(text="NexPlay | Groups Generated")
    await interaction.followup.send(embed=ok_e("Groups Generated!", f"{len(groups)} groups created for **{t['name']}**. Check #{short}-groups."))
    if groups_ch:
        await groups_ch.send(embed=e)


@tree.command(name="results", description="🏅 Post match results for a tournament")
@app_commands.describe(tournament="Tournament name", match_info="Match result (e.g. 'Group A: Team1 2-1 Team2')")
async def cmd_results(interaction: discord.Interaction, tournament: str, match_info: str):
    """Post match results to the results channel and save to DB."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), "results")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    short = t.get("short_name", "")
    results_ch = discord.utils.get(interaction.guild.text_channels, name=f"{short}-results")
    if not results_ch:
        return await interaction.followup.send(embed=err_e(f"#{short}-results channel not found."), ephemeral=True)

    e = discord.Embed(title=f"🏅 Result — {t['name']}", description=match_info, color=0xFFD700, timestamp=datetime.now(timezone.utc))
    e.set_footer(text=f"Posted by {interaction.user.display_name} | NexPlay")
    await results_ch.send(embed=e)
    await interaction.followup.send(embed=ok_e("Result Posted!", f"Match result posted to #{short}-results."), ephemeral=True)


@tree.command(name="export_csv", description="📊 Export tournament registrations as CSV")
@app_commands.describe(tournament="Tournament name")
async def cmd_export_csv(interaction: discord.Interaction, tournament: str):
    """Export registrations as a CSV file."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True, ephemeral=True)
    ok, msg = await check_feature(str(interaction.guild.id), "export_csv")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    regs = await b44_list("Registration", {"tournament_id": t["id"]})
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Team Name", "Captain ID", "Team Members", "Status", "Group", "Seed #", "Logo URL", "Registered At"])
    for r in regs:
        members = r.get("team_members", [])
        if isinstance(members, list): members = ", ".join(members)
        writer.writerow([
            r.get("player_name", ""), r.get("player_discord_id", ""),
            members, r.get("status", ""), r.get("group_label", "—"),
            r.get("seed_number", ""), r.get("logo_url", ""),
            str(r.get("registered_at", ""))[:16],
        ])
    buf.seek(0)
    safe_name = "".join(c for c in t.get("name","export") if c.isalnum() or c in " -_")[:30]
    file = discord.File(io.BytesIO(buf.getvalue().encode()), filename=f"{safe_name}_registrations.csv")
    await interaction.followup.send(embed=ok_e("CSV Export", f"Exported {len(regs)} registrations for **{t['name']}**."), file=file, ephemeral=True)


@tree.command(name="export_registrations", description="📊 Export registrations (alias for /export_csv)")
@app_commands.describe(tournament="Tournament name")
async def cmd_export_registrations(interaction: discord.Interaction, tournament: str):
    """Alias for /export_csv."""
    await cmd_export_csv(interaction, tournament)


@tree.command(name="schedule_post", description="📅 Post the match schedule to the schedule channel")
@app_commands.describe(tournament="Tournament name")
async def cmd_schedule_post(interaction: discord.Interaction, tournament: str):
    """Post tournament schedule to roadmap/results channel."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), "schedule_post")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    short = t.get("short_name", "")
    groups = await b44_list("TournamentGroup", {"tournament_id": t["id"]})
    rounds = int(t.get("rounds", 3))
    date = t.get("tournament_date", "TBD")
    time_str = t.get("tournament_time", "TBD")

    desc = f"📅 **Schedule for {t['name']}**\n🗓️ **Date:** {date} | ⏰ **Time:** {time_str}\n\n"
    for gi, g in enumerate(groups):
        label = g.get("group_label", chr(65+gi))
        names = g.get("player_names", [])
        if isinstance(names, list): names = ", ".join(names)
        desc += f"**Group {label}:** {names}\n"

    desc += "\n"
    stage_names = STAGE_NAMES[:rounds] if rounds <= len(STAGE_NAMES) else [f"Round {i+1}" for i in range(rounds)]
    for i, stage in enumerate(stage_names):
        desc += f"**Stage {i+1}:** {stage} — {date} (Day {i+1})\n"

    e = discord.Embed(title=f"📅 Match Schedule — {t['name']}", description=desc, color=0x1E90FF, timestamp=datetime.now(timezone.utc))
    e.set_footer(text="NexPlay Schedule")

    # Post to roadmap channel
    road_ch = discord.utils.get(interaction.guild.text_channels, name=f"{short}-roadmap")
    if road_ch:
        await road_ch.send(embed=e)

    # Generate schedule GFX (Pro+)
    can_gfx, _ = await check_feature(gid, "ai_gfx_schedule")
    if can_gfx:
        gfx = img_url(t["name"], t.get("game",""), "schedule", f"{len(groups)} groups")
        e.set_image(url=gfx)

    await interaction.followup.send(embed=ok_e("Schedule Posted!", f"Schedule posted to #{short}-roadmap."))


@tree.command(name="roadmap_post", description="🗺️ Post the tournament roadmap")
@app_commands.describe(tournament="Tournament name")
async def cmd_roadmap_post(interaction: discord.Interaction, tournament: str):
    """Post tournament roadmap to the roadmap channel."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), "roadmap_post")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    short = t.get("short_name", "")
    rounds = int(t.get("rounds", 3))
    road_ch = discord.utils.get(interaction.guild.text_channels, name=f"{short}-roadmap")
    if not road_ch:
        return await interaction.followup.send(embed=err_e(f"#{short}-roadmap channel not found."), ephemeral=True)

    stage_names = STAGE_NAMES[:rounds] if rounds <= len(STAGE_NAMES) else [f"Round {i+1}" for i in range(rounds)]
    lines = "\n".join([f"**Stage {i+1}:** {stage_names[i]} — {t.get('tournament_date','TBD')}" for i in range(rounds)])
    e = discord.Embed(title=f"🗺️ {t['name']} — Roadmap", description=lines, color=0xFF6B35, timestamp=datetime.now(timezone.utc))
    e.set_footer(text="NexPlay Roadmap")

    # AI GFX roadmap (Pro+)
    can_gfx, _ = await check_feature(gid, "ai_gfx_poster")
    if can_gfx:
        e.set_image(url=img_url(t["name"], t.get("game",""), "roadmap"))

    await road_ch.send(embed=e)
    await interaction.followup.send(embed=ok_e("Roadmap Posted!", f"Roadmap posted to #{short}-roadmap."))


@tree.command(name="standings_post", description="🏅 Post current standings/leaderboard")
@app_commands.describe(tournament="Tournament name")
async def cmd_standings_post(interaction: discord.Interaction, tournament: str):
    """Post current standings based on match results."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("❌ You need the **Tournament Host** role!"), ephemeral=True)
    await interaction.response.defer(thinking=True)
    ok, msg = await check_feature(str(interaction.guild.id), "standings_post")
    if not ok:
        return await interaction.followup.send(embed=err_e(msg), ephemeral=True)

    gid = str(interaction.guild.id)
    t = await _find_tournament(gid, tournament)
    if not t:
        return await interaction.followup.send(embed=err_e(f"Tournament **{tournament}** not found."), ephemeral=True)

    short = t.get("short_name", "")
    matches = await b44_list("Match", {"tournament_id": t["id"]})
    completed = [m for m in matches if m.get("status") == "completed" and m.get("winner_id")]

    if not completed:
        # No completed matches — show group stage standings instead
        groups = await b44_list("TournamentGroup", {"tournament_id": t["id"]})
        if not groups:
            return await interaction.followup.send(embed=err_e("No matches completed and no groups generated yet. Run /groups first."), ephemeral=True)
        desc = ""
        for g in groups:
            label = g.get("group_label", "?")
            names = g.get("player_names", [])
            if isinstance(names, list): names = "\n".join(f"  {i+1}. {n}" for i, n in enumerate(names))
            desc += f"**🏆 Group {label}**\n{names}\n\n"
        e = discord.Embed(title=f"🏅 Standings — {t['name']} (Group Stage)", description=desc, color=0xFFD700, timestamp=datetime.now(timezone.utc))
    else:
        # Count wins per player
        wins = {}
        for m in completed:
            wid = m.get("winner_id", "")
            wname = m.get("winner_username", "Unknown")
            wins[wname] = wins.get(wname, 0) + 1
        standings = sorted(wins.items(), key=lambda x: -x[1])
        desc = "\n".join([f"{'🥇' if i==0 else '🥈' if i==1 else '🥉' if i==2 else f'{i+1}.'} **{name}** — {w} wins" for i, (name, w) in enumerate(standings)])
        e = discord.Embed(title=f"🏅 Standings — {t['name']}", description=desc, color=0xFFD700, timestamp=datetime.now(timezone.utc))

    e.set_footer(text="NexPlay Standings")
    results_ch = discord.utils.get(interaction.guild.text_channels, name=f"{short}-results")
    if results_ch:
        await results_ch.send(embed=e)
    await interaction.followup.send(embed=ok_e("Standings Posted!", f"Standings posted to #{short}-results."))


@tree.command(name="help", description="📖 Show all NexPlay commands and features")
async def cmd_help(interaction: discord.Interaction):
    """Show help with all available commands based on server's plan."""
    gid = str(interaction.guild.id)
    rec = await get_server_record(gid)
    plan = "trial"
    if rec:
        plan = (rec.get("plan_name") or rec.get("subscription_status") or "trial").lower()
        if plan in ("free trial", "free_trial"): plan = "trial"

    features = PLAN_FEATURES.get(plan, PLAN_FEATURES["trial"])

    e = discord.Embed(title="📖 NexPlay Commands", description=f"**Plan:** {plan.title()} | **Server:** {interaction.guild.name}", color=0x5865F2, timestamp=datetime.now(timezone.utc))

    e.add_field(name="🏆 Tournament", value=(
        "`/create_tournament` — Create a new tournament (3-step form)\n"
        "`/edit_tournament` — Edit tournament details\n"
        "`/register` — Show registration instructions\n"
        "`/groups` — Generate group draws\n"
        "`/results` — Post match results\n"
        "`/standings_post` — Post current standings\n"
        "`/schedule_post` — Post match schedule\n"
        "`/roadmap_post` — Post tournament roadmap\n"
        "`/export_csv` — Export registrations as CSV\n"
        "`/delete_tournament` — Delete tournament + channels\n"
        "`/delete_channel` — Delete a channel\n"
        "`/delete_category` — Delete a category + channels"
    ), inline=False)

    e.add_field(name="🎵 Music", value=(
        "`/play <song>` — Play music from YouTube\n"
        "`/skip` — Skip current song\n"
        "`/queue` — Show music queue\n"
        "`/loop` — Toggle loop\n"
        "`/pause` — Pause music\n"
        "`/resume` — Resume music\n"
        "`/volume <0-100>` — Set volume\n"
        "`/247` — 24/7 auto-play (Gen Z Hindi/Nepali)\n"
        "`/nowplaying` — Now playing\n"
        "`/playlist` — View 24/7 playlist\n"
        "`/leave` — Disconnect bot"
    ), inline=False)

    e.add_field(name="🛠️ Server", value=(
        "`/setup` — Initialize NexPlay channels\n"
        "`/announce` — Post an announcement\n"
        "`/clearlog` — Clear support lock\n"
        "`/help` — This message"
    ), inline=False)

    elite_features = []
    if "ai_support" in features: elite_features.append("🤖 AI Support in #support/#help")
    if "host_game" in features: elite_features.append("🎮 `/host_game` — Mini-games")
    if "suggest_improvement" in features: elite_features.append("💡 `/suggest_improvement` — AI growth advisor")
    if "welcome_message" in features: elite_features.append("👋 Auto welcome messages")
    if "meme_post" in features: elite_features.append("😂 Auto trending memes (15 min)")
    if "daily_report" in features: elite_features.append("📊 Daily Excel reports via DM")
    if "ai_gfx_poster" in features: elite_features.append("🎨 AI GFX posters, groups, results, schedule")

    if elite_features:
        e.add_field(name="✨ Active Features", value="\n".join(elite_features), inline=False)

    e.add_field(name="🔗 Links", value=(
        "Portal: https://nexplay-server-portal.vercel.app\n"
        "Billing: https://nexplay-server-portal.vercel.app/subscription"
    ), inline=False)

    e.set_footer(text="NexPlay Tournament System")
    await interaction.response.send_message(embed=e, ephemeral=True)


@tree.command(name="delete_channel", description="[Host] Delete a channel by name or ID")
@app_commands.describe(channel="The channel to delete (name, ID, or #mention)")
async def cmd_delete_channel(interaction: discord.Interaction, channel: str):
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("You need the **Tournament Host** role to use this command."), ephemeral=True)

    gid = str(interaction.guild.id)
    allowed, reason = await is_allowed(gid, interaction.guild)
    if not allowed:
        return await interaction.response.send_message(embed=err_e(reason), ephemeral=True)

    await interaction.response.defer(thinking=True, ephemeral=True)

    # Parse channel input — #mention, raw ID, or name
    raw = channel.strip().lstrip("#")
    if raw.startswith("<#") and raw.endswith(">"): raw = raw[2:-1]
    ch = interaction.guild.get_channel(int(raw)) if raw.isdigit() else (
        discord.utils.get(interaction.guild.text_channels, name=raw) or
        discord.utils.get(interaction.guild.voice_channels, name=raw))

    if not ch:
        return await interaction.followup.send(embed=err_e(f"Channel `{channel}` not found."), ephemeral=True)

    ch_name = ch.name
    try:
        await ch.delete(reason=f"NexPlay: Deleted by {interaction.user.display_name}")
        await interaction.followup.send(embed=ok_e("Channel Deleted", f"🗑️ Channel **#{ch_name}** has been deleted."), ephemeral=True)
    except discord.Forbidden:
        await interaction.followup.send(embed=err_e("I don't have permission to delete that channel."), ephemeral=True)
    except Exception as e:
        await interaction.followup.send(embed=err_e(f"Failed to delete channel: {e}"), ephemeral=True)


@tree.command(name="delete_category", description="[Host] Delete a category AND all channels inside it")
@app_commands.describe(category="The category to delete (name or ID)")
async def cmd_delete_category(interaction: discord.Interaction, category: str):
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("You need the **Tournament Host** role to use this command."), ephemeral=True)

    gid = str(interaction.guild.id)
    allowed, reason = await is_allowed(gid, interaction.guild)
    if not allowed:
        return await interaction.response.send_message(embed=err_e(reason), ephemeral=True)

    await interaction.response.defer(thinking=True, ephemeral=True)

    # Parse category input
    cat = None
    raw = category.strip()

    if raw.isdigit():
        cat = interaction.guild.get_channel(int(raw))
        if cat and not isinstance(cat, discord.CategoryChannel):
            cat = None
    else:
        cat = discord.utils.get(interaction.guild.categories, name=raw)

    if not cat:
        return await interaction.followup.send(embed=err_e(f"Category `{category}` not found."), ephemeral=True)

    # Delete all channels in category, then the category
    deleted, failed = [], []
    for ch in list(cat.channels):
        try: await ch.delete(reason="NexPlay: Category cleanup"); deleted.append(ch.name)
        except: failed.append(ch.name)
    cat_name = cat.name
    try:
        await cat.delete(reason=f"NexPlay: Deleted by {interaction.user.display_name}")
    except Exception as e:
        return await interaction.followup.send(embed=err_e(f"Deleted {len(deleted)} channels but category failed: {e}"), ephemeral=True)
    desc = f"🗑️ Category **{cat_name}** + **{len(deleted)}** channels deleted."
    if failed: desc += f"\n⚠️ Failed: {', '.join('#'+c for c in failed)}"
    await interaction.followup.send(embed=ok_e("Category Deleted", desc), ephemeral=True)


@tree.command(name="delete_tournament", description="[Host] Delete a tournament: DB record + all channels + category + registrations")
@app_commands.describe(name="The tournament name to delete")
async def cmd_delete_tournament(interaction: discord.Interaction, name: str):
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("You need the **Tournament Host** role to use this command."), ephemeral=True)

    gid = str(interaction.guild.id)
    allowed, reason = await is_allowed(gid, interaction.guild)
    if not allowed:
        return await interaction.response.send_message(embed=err_e(reason), ephemeral=True)

    await interaction.response.defer(thinking=True, ephemeral=True)

    tournament = await _find_tournament(gid, name)
    if not tournament:
        return await interaction.followup.send(embed=err_e(f"Tournament `{name}` not found."), ephemeral=True)

    t_name = tournament.get("name", name)
    t_id = tournament.get("id", "")
    short = tournament.get("short_name", "")
    cat_name = tournament.get("category_name", f"🏆 {short.upper()}" if short else "")

    deleted_channels = []
    failed_channels = []
    category_deleted = False

    # 0. Excel backup BEFORE deleting
    try:
        regs = await b44_list("Registration", {"tournament_id": t_id})
        groups = await b44_list("TournamentGroup", {"tournament_id": t_id})
        matches = await b44_list("Match", {"tournament_id": t_id})
        xl_buf = await build_tournament_export_excel(tournament, regs, groups, matches)
        safe_name = "".join(c for c in t_name if c.isalnum() or c in " -_")[:30].strip() or "tournament"
        filename = f"NexPlay_Backup_{safe_name}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        dm_embed = discord.Embed(title="📦 Tournament Data Backup",
            description=f"Backup of **{t_name}** before deletion.\nRegs: {len(regs)} | Groups: {len(groups)} | Matches: {len(matches)}\nDeleted by: {interaction.user.mention}",
            color=0xFFA500)
        owner = interaction.guild.owner
        try:
            await (await owner.create_dm()).send(embed=dm_embed, file=discord.File(xl_buf, filename=filename))
        except discord.Forbidden:
            await interaction.channel.send(embed=dm_embed, file=discord.File(xl_buf, filename=filename))
    except Exception as e:
        log(f"[Delete] Excel export failed: {e}")

    # 1. Delete channels by short_name prefix
    deleted_channels, failed_channels = [], []
    if short:
        for ch in list(interaction.guild.text_channels):
            if ch.name.lower().startswith(short.lower()):
                try: await ch.delete(reason=f"NexPlay: Tournament '{t_name}' deleted"); deleted_channels.append(ch.name)
                except: failed_channels.append(ch.name)

    # 2. Delete category + remaining channels
    category_deleted = False
    cat = discord.utils.get(interaction.guild.categories, name=cat_name) if cat_name else None
    if cat:
        for ch in list(cat.channels):
            if ch.name not in deleted_channels:
                try: await ch.delete(reason="NexPlay: Tournament cleanup"); deleted_channels.append(ch.name)
                except: failed_channels.append(ch.name)
        try: await cat.delete(reason=f"NexPlay: Tournament '{t_name}' deleted"); category_deleted = True
        except: pass

    # 3-5. Delete all child records from DB
    reg_count = await _b44_delete_by_tournament("Registration", t_id)
    group_count = await _b44_delete_by_tournament("TournamentGroup", t_id)
    match_count = await _b44_delete_by_tournament("Match", t_id)

    # 6. Delete the tournament record itself
    t_deleted = await b44_delete("Tournament", t_id) if t_id else False

    desc = f"🗑️ **{t_name}**\nChannels: {len(deleted_channels)} deleted"
    if failed_channels: desc += f", {len(failed_channels)} failed"
    if category_deleted: desc += f"\nCategory: {cat_name} deleted"
    desc += f"\nRegs: {reg_count} | Groups: {group_count} | Matches: {match_count}"
    desc += f"\nRecord: {'✅ Deleted' if t_deleted else '❌ Failed'}"
    if t_deleted:
        await interaction.followup.send(embed=ok_e("Tournament Deleted", desc), ephemeral=True)
    else:
        await interaction.followup.send(embed=err_e("DB record could not be deleted."), ephemeral=True)


# Per-guild music state
_music_queues: dict[int, deque] = {}
_music_now: dict[int, dict] = {}
_music_loops: dict[int, bool] = {}
_music_volumes: dict[int, float] = {}

# 24/7 auto-play state: guild_id → True/False (bot stays in VC and plays forever)
_music_247: dict[int, bool] = {}
# guild_id → channel_id where 24/7 was activated
_247_channel: dict[int, int] = {}
# guild_id → list of AI-curated songs already played (to avoid repeats)
_247_history: dict[int, list] = {}

# AI-curated Gen Z favorite Hindi & Nepali songs (YouTube search queries)
# These are popular tracks that Gen Z in Nepal/India actually listens to
GENZ_PLAYLIST = [
    # Hindi hits
    "Apna Bana Le Bhediya Arijit Singh",
    "Tum Hi Ho Aashiqui 2 Arijit Singh",
    "Channa Mereya Ae Dil Hai Mushkil",
    "Raabta Agent Vinod Arijit Singh",
    "Kabira Yeh Jawaani Hai Deewani",
    "Phir Bhi Tumko Chaahunga Half Girlfriend",
    "Ik Vaari Aa Raabta Arijit Singh",
    "Hawayein Jab Harry Met Sejal",
    "Tujhe Kitna Chahne Lage Hum Kabir Singh",
    "Pachtaoge Jaani Ve Arijit Singh",
    "Dilbar Dilbar Satyameva Jayate",
    "Kesariya Brahmastra Arijit Singh",
    "Shershaah Raataan Lambiyan",
    "KGF Theme Song Rocky Bhai",
    "Pasoori Ali Sethi Shae Gill",
    "Baarishein Anuv Jain",
    "Alag Aasmaan Anuv Jain",
    "Baatein Ye Kabhi Na Kehna Khamoshiyan",
    "Hamdard Arijit Singh",
    "Duniyaa Luka Chuppi",
    # Nepali hits
    "Saili Ghimire Hemant Sharma",
    "Yeta Timro Kathmandu Maharani",
    "Chyangba Ohoi Pramod Kharel",
    "Laija Re Swaroop Raj Acharya",
    "Pani Pani Rojina Bhandari",
    "Timi Bina Sushant KC",
    "Sarangi Sushant KC",
    "Kabita Khadka Maruni Song",
    "Pritma Lok Dohori Song",
    "Nepali Lok Dohori Latest Hit",
    # Indian pop / Gen Z favorites
    "Brown Munde AP Dhillon",
    "Excuses AP Dhillon",
    "Insane AP Dhillon",
    "Maan Meri Jaan King",
    "Na Ja Pav Dharia",
    "Kina Dekha Indeep Pari",
    "Sufiyana Pyaar Mera Kailash Kher",
    "Bekhayali Kabir Singh",
    "Aabaad Barbaad Arijit Singh",
    "O Bedardeya Tu Jhoothi Main Makkar",
    "Tere Vaal Love Shagun",
    "Heeriye Jasleen Royal Arijit Singh",
    "Tere Pyaar Mein Tu Jhoothi Main Makkar",
    "Pyaar Hota Kayabar Hua Tu Jhoothi Main Makkar",
    "Mast Magan Two States",
    "Gerua Dilwale",
    "Zaalima Raees",
    "Sochta Hoon Kitne Din Bahar Jaani",
    "Lut Gaye Jubin Nautiyal",
    "Barsaat Ki Dhun Jubin Nautiyal",
    "O Aasai Wale Sapne Bula Le",
    "Tera Fitoor Genius Utkarsh",
    "Mere Mehboob Qayamat Ki Raat",
    "Dil Galti Kar Baitha Hai Jubin Nautiyal",
]

def _fmt_dur(sec: int) -> str:
    """Format seconds as m:ss, or LIVE if 0."""
    return f"{sec//60}:{sec%60:02d}" if sec else "LIVE"

def _get_vc(interaction: discord.Interaction):
    """Get the bot's voice client for this guild."""
    return discord.utils.get(bot.voice_clients, guild=interaction.guild)

# yt-dlp options — extract audio only
_YDL_OPTS = {
    "format": "bestaudio/best",
    "quiet": True,
    "no_warnings": True,
    "default_search": "ytsearch",
    "source_address": "0.0.0.0",
    "noplaylist": True,
}

FFMPEG_OPTS = {
    "before_options": "-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5 -vn",
    "options": "-vn",
}


async def _extract_audio(query: str) -> dict:
    """Extract audio info from YouTube URL or search query. Returns dict with title, url, duration, uploader, thumbnail, webpage_url."""
    loop = asyncio.get_event_loop()

    def extract():
        with yt_dlp.YoutubeDL(_YDL_OPTS) as ydl:
            info = ydl.extract_info(query, download=False)
            if "entries" in info:
                info = info["entries"][0]
            return info

    info = await loop.run_in_executor(None, extract)

    # Get the best audio-only stream URL
    audio_url = info.get("url", "")
    if not audio_url:
        for fmt in info.get("formats", []):
            if fmt.get("acodec", "none") != "none" and fmt.get("vcodec", "none") == "none":
                audio_url = fmt["url"]
                break
    if not audio_url and info.get("formats"):
        audio_url = info["formats"][-1]["url"]
    if not audio_url:
        raise RuntimeError("Could not extract audio URL")

    return {
        "title":       info.get("title", "Unknown"),
        "url":         audio_url,
        "webpage_url": info.get("webpage_url", query),
        "duration":    info.get("duration", 0),
        "uploader":    info.get("uploader", "Unknown"),
        "thumbnail":   info.get("thumbnail", ""),
    }


async def _pick_autoplay_song(guild_id: int) -> dict:
    """Pick a random song from the Gen Z playlist that hasn't been played recently."""
    import random
    history = _247_history.get(guild_id, [])
    # If we've played most of the playlist, reset history
    if len(history) > len(GENZ_PLAYLIST) * 0.7:
        history = []
        _247_history[guild_id] = []
    # Pick a song not in recent history
    available = [s for s in GENZ_PLAYLIST if s not in history]
    if not available:
        available = GENZ_PLAYLIST
        _247_history[guild_id] = []
    chosen = random.choice(available)
    _247_history.setdefault(guild_id, []).append(chosen)
    # Extract audio info
    info = await _extract_audio(chosen)
    return {
        "title": info["title"], "url": info["url"], "webpage_url": info["webpage_url"],
        "duration": info["duration"], "uploader": info["uploader"],
        "thumbnail": info["thumbnail"], "requested_by": "🎵 24/7 AutoPlay",
    }

async def _play_next(guild_id: int, voice_client):
    """Play next song from queue, or auto-play from playlist if 24/7 mode is on."""
    queue = _music_queues.get(guild_id)
    is_247 = _music_247.get(guild_id, False)

    # Queue empty + 24/7 on → pick a random song from the curated playlist
    if (not queue or len(queue) == 0) and is_247 and voice_client and voice_client.is_connected():
        try:
            auto_song = await _pick_autoplay_song(guild_id)
            _music_queues.setdefault(guild_id, deque()).append(auto_song)
            queue = _music_queues.get(guild_id)
        except Exception as e:
            print(f"[Music/247] Auto-play pick failed: {e}", flush=True)
            # Retry after a short delay
            await asyncio.sleep(5)
            if is_247 and voice_client.is_connected():
                asyncio.ensure_future(_play_next(guild_id, voice_client))
            return

    if not queue or len(queue) == 0:
        _music_now.pop(guild_id, None)
        return

    song = queue.popleft()
    _music_now[guild_id] = song

    try:
        fresh_info = await _extract_audio(song["webpage_url"])
        audio_url = fresh_info["url"]

        source = discord.FFmpegPCMAudio(audio_url, **FFMPEG_OPTS)
        player = discord.PCMVolumeTransformer(source, volume=_music_volumes.get(guild_id, 0.5))

        def after_play(error):
            if error:
                print(f"[Music] Playback error: {error}", flush=True)
            if _music_loops.get(guild_id):
                _music_queues.setdefault(guild_id, deque()).appendleft(song)
            try:
                asyncio.run_coroutine_threadsafe(
                    _play_next(guild_id, voice_client), bot.loop)
            except Exception as e:
                print(f"[Music] after_play error: {e}", flush=True)

        voice_client.play(player, after=after_play)
        print(f"[Music] Now playing: {song['title']} in guild {guild_id}{' [24/7]' if is_247 else ''}", flush=True)

    except Exception as e:
        print(f"[Music] Error playing next: {e}", flush=True)
        _music_now.pop(guild_id, None)
        # If 24/7 is on, retry with next song after short delay
        if is_247 and voice_client and voice_client.is_connected():
            await asyncio.sleep(3)
            asyncio.ensure_future(_play_next(guild_id, voice_client))
        elif queue:
            await _play_next(guild_id, voice_client)


@tree.command(name="play", description="🎵 Play music from YouTube URL or search query")
async def cmd_play(interaction: discord.Interaction, query: str):
    """Play a song from YouTube URL or search by name."""
    await interaction.response.defer()

    # Check if user is in a voice channel
    if not interaction.user.voice or not interaction.user.voice.channel:
        await interaction.followup.send(embed=err_e("You need to be in a voice channel first!"))
        return

    voice_channel = interaction.user.voice.channel
    guild_id = interaction.guild.id

    # Connect to voice channel
    voice_client = _get_vc(interaction)
    if not voice_client:
        try:
            voice_client = await voice_channel.connect(timeout=30, reconnect=True)
        except Exception as e:
            return await interaction.followup.send(embed=err_e(f"Failed to join voice: {e}"))
    elif voice_client.channel != voice_channel:
        await voice_client.move_to(voice_channel)
    if guild_id not in _music_queues:
        _music_queues[guild_id] = deque()
        _music_volumes[guild_id] = 0.5

    # Extract video info
    try:
        info = await _extract_audio(query)
    except Exception as e:
        await interaction.followup.send(embed=err_e(f"Could not find song: {e}"))
        return

    song = {
        "title":       info["title"],
        "url":         info["url"],
        "webpage_url": info["webpage_url"],
        "duration":    info["duration"],
        "uploader":    info["uploader"],
        "thumbnail":   info["thumbnail"],
        "requested_by": interaction.user.display_name,
    }

    duration_fmt = _fmt_dur(song["duration"])

    # If nothing playing, start immediately
    if not voice_client.is_playing() and not voice_client.is_paused() and len(_music_queues.get(guild_id, [])) == 0:
        _music_queues[guild_id].append(song)
        await interaction.followup.send(embed=ok_e(
            "▶️ Now Playing",
            f"**[{song['title']}]({song['webpage_url']})**\n"
            f"⏱️ {duration_fmt} | 👤 {song['uploader']}\n"
            f"🎵 Requested by {song['requested_by']}"
        ))
        await _play_next(guild_id, voice_client)
    else:
        _music_queues[guild_id].append(song)
        queue_len = len(_music_queues[guild_id])
        await interaction.followup.send(embed=ok_e(
            "➕ Added to Queue",
            f"**[{song['title']}]({song['webpage_url']})**\n"
            f"⏱️ {duration_fmt} | 👤 {song['uploader']}\n"
            f"📋 Position #{queue_len} in queue | Requested by {song['requested_by']}"
        ))


@tree.command(name="skip", description="⏭️ Skip the current song")
async def cmd_skip(interaction: discord.Interaction):
    voice_client = _get_vc(interaction)
    if not voice_client or not voice_client.is_playing():
        await interaction.response.send_message(embed=err_e("Nothing is playing right now."), ephemeral=True)
        return

    current = _music_now.get(interaction.guild.id, {})
    title = current.get("title", "Unknown")
    _music_loops[interaction.guild.id] = False  # reset loop on skip
    voice_client.stop()  # triggers after_play → plays next
    await interaction.response.send_message(embed=ok_e("⏭️ Skipped", f"**{title}** skipped."))


@tree.command(name="stop", description="⏹️ Stop music, clear queue, disable 24/7 mode")
async def cmd_stop(interaction: discord.Interaction):
    voice_client = _get_vc(interaction)
    guild_id = interaction.guild.id
    was_247 = _music_247.get(guild_id, False)

    _music_queues[guild_id] = deque()
    _music_now.pop(guild_id, None)
    _music_loops[guild_id] = False
    _music_247[guild_id] = False  # Disable 24/7
    _247_channel.pop(guild_id, None)

    if voice_client:
        if voice_client.is_playing():
            voice_client.stop()
        await voice_client.disconnect()
    msg = "Music stopped, queue cleared, 24/7 disabled." if was_247 else "Music stopped and queue cleared."
    await interaction.response.send_message(embed=ok_e("⏹️ Stopped", msg))


@tree.command(name="queue", description="📋 Show the current music queue")
async def cmd_queue(interaction: discord.Interaction):
    guild_id = interaction.guild.id
    queue = _music_queues.get(guild_id, deque())
    now = _music_now.get(guild_id)

    if not now and len(queue) == 0:
        await interaction.response.send_message(embed=err_e("The queue is empty."))
        return

    desc = ""
    if now:
        dur = now.get("duration", 0)
        dur_fmt = _fmt_dur(dur)
        desc += f"▶️ **NOW PLAYING:** [{now['title']}]({now.get('webpage_url', '')})\n⏱️ {dur_fmt} | 👤 {now.get('uploader', '')}\n\n"
    if queue:
        desc += "**UP NEXT:**\n"
        for i, song in enumerate(list(queue)[:10], 1):
            dur = song.get("duration", 0)
            dur_fmt = _fmt_dur(dur)
            desc += f"`{i}.` [{song['title']}]({song.get('webpage_url', '')}) — {dur_fmt}\n"
        if len(queue) > 10:
            desc += f"\n*...and {len(queue) - 10} more*\n"
    else:
        desc += "*No songs in queue.*"

    if _music_loops.get(guild_id):
        desc += "\n🔁 **Loop mode: ON**"
    if _music_247.get(guild_id):
        desc += "\n♾️ **24/7 AutoPlay: ON**"

    embed = discord.Embed(title="🎵 Music Queue", description=desc, color=0x1DB954)
    await interaction.response.send_message(embed=embed)


@tree.command(name="loop", description="🔁 Toggle loop for the current song")
async def cmd_loop(interaction: discord.Interaction):
    guild_id = interaction.guild.id
    current = _music_loops.get(guild_id, False)
    _music_loops[guild_id] = not current
    status = "ON 🔁" if not current else "OFF"
    await interaction.response.send_message(embed=ok_e("🔁 Loop", f"Loop mode is now **{status}**."))


@tree.command(name="leave", description="👋 Disconnect bot, disable 24/7 mode")
async def cmd_leave(interaction: discord.Interaction):
    voice_client = _get_vc(interaction)
    if not voice_client:
        return await interaction.response.send_message(embed=err_e("I'm not in a voice channel."))
    gid = interaction.guild.id
    _music_queues[gid] = deque()
    _music_now.pop(gid, None)
    _music_247[gid] = False
    _247_channel.pop(gid, None)
    await voice_client.disconnect()
    await interaction.response.send_message(embed=ok_e("👋 Left", "Disconnected. 24/7 disabled."))


@tree.command(name="volume", description="🔊 Set music volume (0-100)")
async def cmd_volume(interaction: discord.Interaction, level: int):
    if level < 0 or level > 100:
        await interaction.response.send_message(embed=err_e("Volume must be between 0 and 100."))
        return
    vol = level / 100
    _music_volumes[interaction.guild.id] = vol
    voice_client = _get_vc(interaction)
    if voice_client and hasattr(voice_client.source, "volume"):
        voice_client.source.volume = vol
    await interaction.response.send_message(embed=ok_e("🔊 Volume", f"Volume set to **{level}%**."))


@tree.command(name="pause", description="⏸️ Pause the current song")
async def cmd_pause(interaction: discord.Interaction):
    voice_client = _get_vc(interaction)
    if not voice_client or not voice_client.is_playing():
        await interaction.response.send_message(embed=err_e("Nothing is playing right now."), ephemeral=True)
        return
    voice_client.pause()
    await interaction.response.send_message(embed=ok_e("⏸️ Paused", "Music paused. Use /play to resume."))


@tree.command(name="resume", description="▶️ Resume the paused song")
async def cmd_resume(interaction: discord.Interaction):
    voice_client = _get_vc(interaction)
    if not voice_client or not voice_client.is_paused():
        await interaction.response.send_message(embed=err_e("Nothing is paused right now."), ephemeral=True)
        return
    voice_client.resume()
    await interaction.response.send_message(embed=ok_e("▶️ Resumed", "Music resumed."))


# ──────────────────────────────────────────────────────────
#  24/7 AUTO-PLAY SYSTEM
# ──────────────────────────────────────────────────────────

@tree.command(name="247", description="🎵 Enable 24/7 music — bot stays in VC and plays forever")
async def cmd_247(interaction: discord.Interaction):
    """Enable 24/7 auto-play mode. Bot joins the user's voice channel and
    continuously plays random Gen Z Hindi/Nepali songs from the curated playlist."""
    if not is_staff(interaction.user):
        return await interaction.response.send_message(embed=err_e("Staff only to enable 24/7 mode."), ephemeral=True)

    gid = interaction.guild.id

    # If already in 24/7 mode, disable it
    if _music_247.get(gid, False):
        _music_247[gid] = False
        _247_channel.pop(gid, None)
        voice_client = _get_vc(interaction)
        if voice_client and voice_client.is_playing():
            voice_client.stop()
        # Don't disconnect — let the current queue finish or user can /leave
        return await interaction.response.send_message(embed=ok_e("🎵 24/7 Disabled",
            "Auto-play stopped. Bot will disconnect when queue ends.\nUse `/leave` to disconnect now."))

    # User must be in a voice channel
    if not interaction.user.voice or not interaction.user.voice.channel:
        return await interaction.response.send_message(embed=err_e("Join a voice channel first, then run `/247`."))

    voice_channel = interaction.user.voice.channel
    voice_client = _get_vc(interaction)

    if not voice_client:
        try:
            voice_client = await voice_channel.connect(timeout=30, reconnect=True)
        except Exception as e:
            return await interaction.response.send_message(embed=err_e(f"Failed to join voice: {e}"))
    elif voice_client.channel != voice_channel:
        await voice_client.move_to(voice_channel)

    # Enable 24/7
    _music_247[gid] = True
    _247_channel[gid] = voice_channel.id
    if gid not in _music_queues:
        _music_queues[gid] = deque()
        _music_volumes[gid] = 0.5

    await interaction.response.send_message(embed=ok_e("🎵 24/7 AutoPlay Enabled!",
        f"Bot is now in **{voice_channel.name}** and will play music 24/7!\n\n"
        f"🎶 Playlist: **Gen Z Hindi & Nepali Hits** ({len(GENZ_PLAYLIST)} songs)\n"
        f"🤖 AI-curated: Arijit Singh, AP Dhillon, Sushant KC, Jubin Nautiyal & more\n"
        f"♾️ Random shuffle — no repeats for 70% of playlist\n\n"
        f"**Controls:**\n"
        f"• `/skip` — skip current song\n"
        f"• `/play <song>` — add a song (plays after current)\n"
        f"• `/queue` — see what's playing\n"
        f"• `/stop` — stop everything & disconnect\n"
        f"• `/247` again — toggle off\n"
        f"• `/playlist` — view full playlist"))

    # If nothing is playing, start auto-play
    if not voice_client.is_playing() and not voice_client.is_paused():
        await _play_next(gid, voice_client)


@tree.command(name="autoplay", description="🎵 Toggle auto-play (same as /247)")
async def cmd_autoplay(interaction: discord.Interaction):
    """Alias for /247."""
    await cmd_247(interaction)


@tree.command(name="playlist", description="📜 View the 24/7 auto-play playlist")
async def cmd_playlist(interaction: discord.Interaction):
    """Show the curated Gen Z Hindi/Nepali playlist."""
    desc = "**🎶 Gen Z Hindi & Nepali Hits**\n\n"
    desc += "**🇮🇳 Hindi Hits:**\n"
    hindi_songs = [s for s in GENZ_PLAYLIST if "Arijit" in s or "Jubin" in s or "KGF" in s
                   or "Brahmastra" in s or "Shershaah" in s or "Dhillon" in s or "King" in s
                   or "Anuv" in s or "Kher" in s or "Nautiyal" in s or "Jaani" in s
                   or "Raabta" in s or "Dilwale" in s or "Raees" in s or "Genius" in s]
    nepali_songs = [s for s in GENZ_PLAYLIST if s not in hindi_songs]

    for i, s in enumerate(hindi_songs[:20], 1):
        desc += f"`{i:2}.` {s}\n"
    if len(hindi_songs) > 20:
        desc += f"*...and {len(hindi_songs)-20} more Hindi tracks*\n"

    desc += f"\n**🇳🇵 Nepali Hits:**\n"
    for i, s in enumerate(nepali_songs, 1):
        desc += f"`{i:2}.` {s}\n"

    desc += f"\n**Total: {len(GENZ_PLAYLIST)} songs** — random shuffle in 24/7 mode\n"
    desc += "Use `/247` to start auto-play! 🎵"

    # Split into multiple embeds if too long
    if len(desc) > 4000:
        desc = desc[:3990] + "..."

    e = discord.Embed(title="📜 NexPlay 24/7 Playlist", description=desc, color=0x1DB954,
        timestamp=datetime.now(timezone.utc))
    e.set_footer(text="AI-curated Gen Z Hindi & Nepali playlist")
    await interaction.response.send_message(embed=e)


@tree.command(name="nowplaying", description="🎵 Show what's currently playing")
async def cmd_nowplaying(interaction: discord.Interaction):
    """Show the currently playing song with details."""
    gid = interaction.guild.id
    now = _music_now.get(gid)
    if not now:
        return await interaction.response.send_message(embed=err_e("Nothing is playing right now."))

    dur = _fmt_dur(now.get("duration", 0))
    is_247 = _music_247.get(gid, False)
    requester = now.get("requested_by", "Unknown")

    e = discord.Embed(title="🎵 Now Playing", color=0x1DB954, timestamp=datetime.now(timezone.utc))
    e.description = f"**[{now['title']}]({now.get('webpage_url', '')})**\n⏱️ {dur} | 👤 {now.get('uploader', '')}\n🎵 Requested by {requester}"
    if is_247:
        e.description += "\n♾️ **24/7 AutoPlay is ON**"
    if now.get("thumbnail"):
        e.set_thumbnail(url=now["thumbnail"])
    e.set_footer(text="NexPlay Music")
    await interaction.response.send_message(embed=e)


async def _247_health_check():
    """Background task: every 60s, check all guilds with 24/7 enabled.
    If bot disconnected from VC, reconnect and resume auto-play."""
    await bot.wait_until_ready()
    print("[NexPlay] 24/7 health check started.", flush=True)
    while not bot.is_closed():
        try:
            for guild in bot.guilds:
                gid = guild.id
                if not _music_247.get(gid, False):
                    continue
                vc = discord.utils.get(bot.voice_clients, guild=guild)
                ch_id = _247_channel.get(gid)
                # Bot not connected but 24/7 is on → reconnect
                if not vc or not vc.is_connected():
                    ch = guild.get_channel(ch_id) if ch_id else None
                    if not ch:
                        # Find any voice channel the bot can join
                        ch = discord.utils.find(lambda c: c.permissions_for(guild.me).connect, guild.voice_channels)
                    if ch:
                        try:
                            vc = await ch.connect(timeout=15, reconnect=True)
                            _247_channel[gid] = ch.id
                            print(f"[NexPlay/247] Reconnected to {ch.name} in {guild.name}", flush=True)
                            if not vc.is_playing():
                                await _play_next(gid, vc)
                        except Exception as e:
                            print(f"[NexPlay/247] Reconnect failed in {guild.name}: {e}", flush=True)
                # Connected but nothing playing → resume auto-play
                elif vc.is_connected() and not vc.is_playing() and not vc.is_paused():
                    if len(_music_queues.get(gid, deque())) == 0:
                        await _play_next(gid, vc)
        except Exception as e:
            print(f"[NexPlay/247] Health check error: {e}", flush=True)
        await asyncio.sleep(60)


@bot.event
async def on_guild_join(guild: discord.Guild):
    """Auto-register new servers OR reactivate previously removed servers.
    When bot is re-invited to a server it was previously in, it restores the
    old server record (plan, tournament count, etc.) instead of creating a new one."""
    print(f"[NexPlay] 🆕 Joined: {guild.name} ({guild.id}) — {guild.member_count} members", flush=True)
    gid = str(guild.id)
    existing = await b44_list("Server", {"guild_id": gid})
    if not existing:
        # Brand new server — register as Free Trial
        await register_server(guild)
        print(f"[NexPlay] ✅ Registered {guild.name} as Free Trial (3 tournaments)", flush=True)
    else:
        srv = existing[0]
        status = srv.get("subscription_status", "trial")
        if status == "inactive":
            # Server was previously removed — reactivate it, preserving plan + data
            await b44_update("Server", srv["id"], {
                "subscription_status": "trial" if srv.get("plan_name", "Free Trial") == "Free Trial" else "active",
                "guild_name": guild.name,  # Update name in case it changed
                "owner_id": str(guild.owner.id) if guild.owner else srv.get("owner_id", ""),
                "owner_name": guild.owner.display_name if guild.owner else srv.get("owner_name", ""),
                "member_count": guild.member_count or 0,
                "last_active": now_iso(),
            })
            plan = srv.get("plan_name", "Free Trial")
            t_used = srv.get("tournaments_used", 0)
            print(f"[NexPlay] ♻️ Reactivated {guild.name} — Plan: {plan}, Tournaments used: {t_used}", flush=True)
        else:
            print(f"[NexPlay] {guild.name} already registered (status: {status})", flush=True)

    # Auto-create needed roles if missing
    try:
        for role_name in ("Tournament Host", "NexPlay Admin"):
            if not discord.utils.get(guild.roles, name=role_name):
                await guild.create_role(
                    name=role_name,
                    colour=discord.Colour.gold() if "Admin" in role_name else discord.Colour.blue(),
                    mentionable=True,
                    reason="NexPlay: Auto-created role for tournament management"
                )
                print(f"[NexPlay] Created role '{role_name}' in {guild.name}", flush=True)
    except Exception as e:
        print(f"[NexPlay] Role creation failed in {guild.name}: {e}", flush=True)

    # Welcome message
    for ch in guild.text_channels:
        if ch.permissions_for(guild.me).send_messages:
            await ch.send(embed=ok_e("👋 NexPlay is here!",
                f"Welcome **{guild.name}**! 🎮\n"
                "🏆 `/create_tournament` · ✍️ `/register` · 🎨 GFX · 🤖 AI support\n"
                "Run `/setup` to get started. Free Trial = 3 tournaments.\n"
                "Portal: https://nexplay-server-portal.vercel.app"))
            break

    await _update_member_count(guild)

@bot.event
async def on_guild_remove(guild: discord.Guild):
    """Bot removed from a server — mark inactive but DON'T delete data."""
    print(f"[NexPlay] ❌ Removed from: {guild.name} ({guild.id})", flush=True)
    srv = await get_server_record(str(guild.id))
    if srv:
        await b44_update("Server", srv["id"], {
            "subscription_status": "inactive",
            "last_active": now_iso(),
        })
        print(f"[NexPlay] Marked {guild.name} as inactive (data preserved)", flush=True)
    # Clean up meme cache + 24/7 state
    _last_meme_url.pop(guild.id, None)
    _meme_channel_cache.pop(guild.id, None)
    _music_247.pop(guild.id, None)
    _247_channel.pop(guild.id, None)
    _247_history.pop(guild.id, None)


@bot.event
async def on_member_remove(member: discord.Member):
    """Update member count when someone leaves."""
    await _update_member_count(member.guild)


bot.run(BOT_TOKEN)
